from __future__ import annotations

import importlib
import json
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook


class _FakeCell:
    def __init__(self, value: str, coordinate: str) -> None:
        self.value = value
        self.coordinate = coordinate


class _StreamingSheet:
    title = "运营日报"
    max_row = 2
    max_column = 2

    def iter_rows(self, *args, **kwargs):
        values_only = kwargs.get("values_only", False)
        rows = [("日期", "销售额"), ("2026-05-01", 100)]
        if values_only:
            yield from rows
            return
        for row_index, row in enumerate(rows, start=1):
            yield tuple(_FakeCell(str(value), f"R{row_index}C{column_index}") for column_index, value in enumerate(row, start=1))


class _Workbook:
    def __init__(self) -> None:
        self.worksheets = [_StreamingSheet()]
        self.sheetnames = ["运营日报"]

    def __getitem__(self, sheet_name: str):
        if sheet_name != "运营日报":
            raise KeyError(sheet_name)
        return self.worksheets[0]


class TableCleaningMediumExcelTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.raw_dir = self.root / "raw"
        self.data_dir = self.root / "data"
        self.cleaned_dir = self.data_dir / "cleaned"
        self.wiki_dir = self.root / "wiki"
        self.raw_dir.mkdir(parents=True)
        self.cleaned_dir.mkdir(parents=True)
        self.wiki_dir.mkdir(parents=True)
        os.environ["A2A_RAW_DIR"] = str(self.raw_dir)
        os.environ["A2A_DATA_DIR"] = str(self.data_dir)
        os.environ["A2A_CLEANED_DIR"] = str(self.cleaned_dir)
        os.environ["A2A_WIKI_DIR"] = str(self.wiki_dir)

        import src.a2a_ecommerce_demo.table_cleaning_tools as table_cleaning_tools

        self.table_cleaning_tools = importlib.reload(table_cleaning_tools)

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    def _write_medium_workbook_placeholder(self) -> None:
        path = self.raw_dir / "1、TM - UNOVE运营日报.xlsx"
        with path.open("wb") as file:
            file.truncate(20 * 1024 * 1024)

    def test_medium_workbooks_use_streaming_mode_for_profile_and_cleaning(self) -> None:
        self._write_medium_workbook_placeholder()

        with patch("openpyxl.load_workbook", side_effect=lambda *args, **kwargs: _Workbook()) as load_workbook:
            profile = json.loads(self.table_cleaning_tools.profile_excel_file("1、TM - UNOVE运营日报.xlsx"))
            clean = json.loads(self.table_cleaning_tools.clean_excel_to_csv("1、TM - UNOVE运营日报.xlsx"))

        self.assertTrue(profile["large_file_safe_mode"])
        self.assertTrue(clean["results"][0]["large_file_safe_mode"])
        self.assertEqual(2, load_workbook.call_count)
        self.assertTrue(all(call.kwargs["read_only"] for call in load_workbook.call_args_list))
        self.assertTrue(all(call.kwargs["data_only"] for call in load_workbook.call_args_list))

    def test_batch_cleaning_skips_medium_workbooks_for_large_file_pipeline(self) -> None:
        self._write_medium_workbook_placeholder()

        with patch("openpyxl.load_workbook", side_effect=AssertionError("medium workbook should be skipped")):
            result = json.loads(self.table_cleaning_tools.clean_all_excel_files())

        self.assertEqual(1, result["processed"])
        self.assertTrue(result["results"][0]["skipped_for_large_file_pipeline"])

    def test_cleaning_falls_back_to_ooxml_when_openpyxl_rejects_styles(self) -> None:
        path = self.raw_dir / "GMV-渠道部品牌销售数据统计.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "GMV"
        sheet.append(["日期", "品牌", "GMV"])
        sheet.append(["2026-05-01", "UNOVE", 1200])
        workbook.save(path)

        with patch("openpyxl.load_workbook", side_effect=TypeError("expected <class 'openpyxl.styles.fills.Fill'>")):
            result = json.loads(self.table_cleaning_tools.clean_excel_to_csv(path.name))

        output_csv = Path(result["results"][0]["output_csv"])
        self.assertTrue(output_csv.exists())
        self.assertEqual("ooxml", result["results"][0]["scan"]["fallback"])
        self.assertIn("UNOVE", output_csv.read_text(encoding="utf-8-sig"))


if __name__ == "__main__":
    unittest.main()
