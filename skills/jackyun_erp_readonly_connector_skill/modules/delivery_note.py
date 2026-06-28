"""
Excel template delivery-note generation.

This module generalizes the learned "申通仓出库单" task:
- Query sales orders with goodsDetail fields.
- Fill only the template cells/columns that the configured warehouse template needs.
- Respect merged cells by writing to the merged area's top-left cell.
- Keep display cells separate from file-name data.
"""
from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from jackyun_api import JackyunValidationError
from modules.sales_order import DEFAULT_TRADE_FIELDS, query_trade_by_no


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_TEMPLATE_PATH = PROJECT_ROOT / "templates" / "shentong_delivery_note.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "output"

DELIVERY_NOTE_TRADE_FIELDS = (
    DEFAULT_TRADE_FIELDS
    + ",goodsDetail.goodsNo,goodsDetail.goodsName,goodsDetail.barcode,goodsDetail.sellCount"
)

SHENTONG_TEMPLATE_CONFIG = {
    "sheet_name": None,
    "order_cell": "F5",
    "order_label": "销售单/采购单：",
    "line_start_row": 10,
    "line_end_row": 40,
    "summary_label_cell": "A41",
    "quantity_summary_cell": "F41",
    "quantity_summary_column": "F",
    "columns": {
        "sequence": "A",
        "barcode": "B",
        "goods_name": "E",
        "quantity": "F",
    },
    "blank_columns": ["C", "D", "I"],
    "copy_style_from_row": 40,
}


def _as_list(value: Any) -> list:
    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        return [value]
    return []


def _safe_filename_part(value: str) -> str:
    text = str(value or "").strip()
    return re.sub(r'[\\/:*?"<>|\\n\\r]+', "_", text) or "unknown"


def _trade_date_code(trade_no: str) -> str:
    match = re.search(r"(\d{6})", str(trade_no or ""))
    return match.group(1) if match else ""


def _default_output_path(trade_no: str, output_dir: str | Path = None) -> Path:
    target_dir = Path(output_dir) if output_dir else DEFAULT_OUTPUT_DIR
    date_code = _trade_date_code(trade_no)
    prefix = f"{date_code}订单出库单" if date_code else "订单出库单"
    return target_dir / f"{prefix}（{_safe_filename_part(trade_no)}）.xlsx"


def merged_top_left(ws: Worksheet, coordinate: str) -> str:
    """
    Return the writable coordinate for a cell, respecting merged ranges.
    """
    cell = ws[coordinate]
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range.start_cell.coordinate
    return cell.coordinate


def set_template_cell(ws: Worksheet, coordinate: str, value: Any):
    ws[merged_top_left(ws, coordinate)] = value


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int):
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _ensure_line_capacity(ws: Worksheet, goods_count: int, config: dict) -> int:
    start = int(config["line_start_row"])
    end = int(config["line_end_row"])
    capacity = end - start + 1
    if goods_count <= capacity:
        return end + 1

    overflow = goods_count - capacity
    summary_row = end + 1
    ws.insert_rows(summary_row, overflow)
    style_row = int(config.get("copy_style_from_row") or end)
    for row in range(summary_row, summary_row + overflow):
        _copy_row_style(ws, style_row, row)
    return summary_row + overflow


def _goods_detail_rows(trade: dict) -> list[dict]:
    goods = _as_list(trade.get("goodsDetail"))
    if goods:
        return goods
    # Some API variants nest details under tradeOrder; keep tolerant.
    return _as_list((trade.get("tradeOrder") or {}).get("goodsDetail"))


def _quantity(value: Any) -> Any:
    if value in (None, ""):
        return 0
    try:
        number = float(value)
        return int(number) if number.is_integer() else number
    except (TypeError, ValueError):
        return value


def _fill_goods_lines(ws: Worksheet, goods_list: list[dict], config: dict) -> int:
    start = int(config["line_start_row"])
    summary_row = _ensure_line_capacity(ws, len(goods_list), config)
    columns = config["columns"]
    blank_columns = config.get("blank_columns") or []

    for offset, item in enumerate(goods_list):
        row = start + offset
        ws[f"{columns['sequence']}{row}"] = offset + 1
        ws[f"{columns['barcode']}{row}"] = item.get("barcode") or item.get("skuBarcode") or item.get("goodsNo") or ""
        ws[f"{columns['goods_name']}{row}"] = item.get("goodsName") or item.get("skuName") or ""
        ws[f"{columns['quantity']}{row}"] = _quantity(item.get("sellCount") or item.get("sendCount") or item.get("quantity"))
        for column in blank_columns:
            ws[f"{column}{row}"] = None

    return summary_row


def generate_delivery_note_from_trade(
    trade_no: str,
    template_path: str | Path = None,
    output_path: str | Path = None,
    output_dir: str | Path = None,
    config: dict = None,
) -> dict:
    """
    Generate an Excel delivery note from a sales order.

    The default profile is the learned Shentong warehouse template, but callers
    may pass another template/config for other warehouse forms.
    """
    if not str(trade_no or "").strip():
        raise JackyunValidationError("生成模板出库单必须提供销售单号 trade_no")

    template = Path(template_path) if template_path else DEFAULT_TEMPLATE_PATH
    if not template.exists():
        raise FileNotFoundError(f"模板文件不存在: {template}")

    profile = {**SHENTONG_TEMPLATE_CONFIG, **(config or {})}
    trade = query_trade_by_no(trade_no, fields=DELIVERY_NOTE_TRADE_FIELDS)
    if not trade:
        raise JackyunValidationError(f"未找到销售单: {trade_no}")

    goods_list = _goods_detail_rows(trade)
    if not goods_list:
        raise JackyunValidationError(
            f"销售单 {trade_no} 未返回货品明细，请确认查询 fields 包含 goodsDetail.goodsName/barcode/sellCount"
        )

    wb = load_workbook(template)
    ws = wb[profile["sheet_name"]] if profile.get("sheet_name") else wb[wb.sheetnames[0]]

    order_text = f"{profile.get('order_label', '')}{trade_no}"
    set_template_cell(ws, profile["order_cell"], order_text)
    summary_row = _fill_goods_lines(ws, goods_list, profile)
    summary_column = profile.get("quantity_summary_column", "F")
    summary_cell = f"{summary_column}{summary_row}"
    ws[summary_cell] = f"=SUM({summary_column}{profile['line_start_row']}:{summary_column}{summary_row - 1})"

    final_output = Path(output_path) if output_path else _default_output_path(trade_no, output_dir=output_dir)
    final_output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(final_output)

    return {
        "output_path": str(final_output),
        "template_path": str(template),
        "trade_no": trade_no,
        "goods_count": len(goods_list),
        "summary_row": summary_row,
        "summary_cell": summary_cell,
        "filled_columns": profile["columns"],
        "blank_columns": profile.get("blank_columns", []),
        "fields": DELIVERY_NOTE_TRADE_FIELDS,
        "notes": [
            "销售单号写入模板展示单元格，但文件名使用原始 trade_no，避免展示文字污染文件名",
            "合并单元格只写入合并区域左上角",
            "默认申通模板只填序号、条码、品名、数量；包装规格/总箱数/出入库仓库保持空白",
        ],
    }
