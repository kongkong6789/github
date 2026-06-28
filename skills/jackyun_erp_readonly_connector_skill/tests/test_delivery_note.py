import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.delivery_note import generate_delivery_note_from_trade, merged_top_left


def _make_template(path: str):
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("F5:J7")
    ws["F5"] = "销售单/采购单："
    ws["A9"] = "序號"
    ws["B9"] = "商品條碼"
    ws["C9"] = "包裝規格"
    ws["D9"] = "總箱數"
    ws["E9"] = "產品品名"
    ws["F9"] = "數量"
    ws["I9"] = "出/入库仓库"
    ws["A41"] = "滙總:"
    ws["F41"] = "=SUM(F10:F40)"
    wb.save(path)


class TestDeliveryNoteGeneration(unittest.TestCase):
    def test_merged_top_left(self):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("F5:J7")
        self.assertEqual(merged_top_left(ws, "J7"), "F5")
        self.assertEqual(merged_top_left(ws, "A1"), "A1")

    @patch("modules.delivery_note.query_trade_by_no")
    def test_generate_shentong_template_keeps_blank_columns_and_summary(self, mock_query):
        mock_query.return_value = {
            "tradeNo": "YR260429001022",
            "goodsDetail": [
                {"barcode": "BAR001", "goodsName": "货品A", "sellCount": "2"},
                {"goodsNo": "G002", "goodsName": "货品B", "sellCount": "3"},
            ],
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            template_path = os.path.join(tmp_dir, "template.xlsx")
            output_path = os.path.join(tmp_dir, "out.xlsx")
            _make_template(template_path)

            result = generate_delivery_note_from_trade(
                "YR260429001022",
                template_path=template_path,
                output_path=output_path,
            )

            wb = load_workbook(output_path, data_only=False)
            ws = wb.active

        self.assertEqual(result["goods_count"], 2)
        self.assertEqual(result["summary_cell"], "F41")
        self.assertEqual(ws["F5"].value, "销售单/采购单：YR260429001022")
        self.assertEqual(ws["A10"].value, 1)
        self.assertEqual(ws["B10"].value, "BAR001")
        self.assertEqual(ws["E10"].value, "货品A")
        self.assertEqual(ws["F10"].value, 2)
        self.assertEqual(ws["B11"].value, "G002")
        self.assertIsNone(ws["C10"].value)
        self.assertIsNone(ws["D10"].value)
        self.assertIsNone(ws["I10"].value)
        self.assertEqual(ws["F41"].value, "=SUM(F10:F40)")
        self.assertIsNone(ws["D41"].value)

    @patch("modules.delivery_note.query_trade_by_no")
    def test_generate_inserts_rows_for_overflow_and_uses_trade_no_filename(self, mock_query):
        mock_query.return_value = {
            "tradeNo": "YR260429001022",
            "goodsDetail": [
                {"barcode": f"BAR{i}", "goodsName": f"货品{i}", "sellCount": "1"}
                for i in range(3)
            ],
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            template_path = os.path.join(tmp_dir, "template.xlsx")
            _make_template(template_path)

            result = generate_delivery_note_from_trade(
                "YR260429001022",
                template_path=template_path,
                output_dir=tmp_dir,
                config={
                    "line_start_row": 10,
                    "line_end_row": 11,
                    "quantity_summary_cell": "F12",
                    "copy_style_from_row": 11,
                },
            )

            output_path = Path(result["output_path"])
            wb = load_workbook(output_path, data_only=False)
            ws = wb.active

        self.assertEqual(output_path.name, "260429订单出库单（YR260429001022）.xlsx")
        self.assertEqual(result["summary_row"], 13)
        self.assertEqual(ws["F13"].value, "=SUM(F10:F12)")
        self.assertEqual(ws["B12"].value, "BAR2")


if __name__ == "__main__":
    unittest.main()
