from __future__ import annotations

import csv
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from src.a2a_ecommerce_demo.xlsx_fallback_tools import read_xlsx_sheets

PROJECT_ROOT = Path(__file__).resolve().parents[2]
RAW_DIR = Path(os.getenv("A2A_RAW_DIR", PROJECT_ROOT / "raw")).resolve()
DATA_DIR = Path(os.getenv("A2A_DATA_DIR", PROJECT_ROOT / "data")).resolve()
CLEANED_DIR = Path(os.getenv("A2A_CLEANED_DIR", DATA_DIR / "cleaned")).resolve()
WIKI_DIR = Path(os.getenv("A2A_WIKI_DIR", PROJECT_ROOT / "wiki")).resolve()

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
LARGE_EXCEL_BYTES = int(os.getenv("A2A_INTERACTIVE_EXCEL_BYTES", str(10 * 1024 * 1024)))
HUGE_EXCEL_BYTES = 100 * 1024 * 1024
DEFAULT_LARGE_CLEAN_ROW_LIMIT = 50000


def _ensure_dirs() -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    CLEANED_DIR.mkdir(parents=True, exist_ok=True)
    WIKI_DIR.mkdir(parents=True, exist_ok=True)


def _safe_under(path: Path, roots: list[Path]) -> Path:
    resolved = path.resolve()
    if not any(root in [resolved, *resolved.parents] for root in roots):
        raise ValueError(f"Refusing to access outside allowed directories: {resolved}")
    return resolved


def _raw_path(path: str) -> Path:
    target = Path(path)
    if not target.is_absolute():
        if target.parts and target.parts[0].lower() == "raw":
            target = Path(*target.parts[1:])
        target = RAW_DIR / target
    target = _safe_under(target, [RAW_DIR])
    if target.suffix.lower() not in EXCEL_SUFFIXES:
        raise ValueError(f"Only .xlsx/.xlsm files are supported for Excel cleaning: {target.name}")
    if not target.exists():
        raise FileNotFoundError(f"Raw Excel file not found: {target}")
    return target


def _is_large_excel(path: Path) -> bool:
    return path.stat().st_size >= LARGE_EXCEL_BYTES


def _is_huge_excel(path: Path) -> bool:
    return path.stat().st_size >= HUGE_EXCEL_BYTES


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _normalize_header(value: str, fallback_index: int, used: dict[str, int]) -> str:
    cleaned = re.sub(r"\s+", "_", value.strip())
    cleaned = re.sub(r"[^\w\u4e00-\u9fff]+", "_", cleaned, flags=re.UNICODE).strip("_")
    cleaned = cleaned[:60] or f"column_{fallback_index}"
    count = used.get(cleaned, 0)
    used[cleaned] = count + 1
    if count:
        return f"{cleaned}_{count + 1}"
    return cleaned


def _non_empty_count(values: list[str]) -> int:
    return sum(1 for value in values if value)


def _looks_numeric(value: str) -> bool:
    if not value:
        return False
    try:
        float(value.replace(",", ""))
        return True
    except ValueError:
        return False


def _row_score(values: list[str]) -> float:
    non_empty = _non_empty_count(values)
    if non_empty < 2:
        return 0.0
    text_count = sum(1 for value in values if value and not _looks_numeric(value))
    unique_count = len({value for value in values if value})
    duplicate_penalty = max(0, non_empty - unique_count) * 0.2
    return non_empty + text_count * 0.7 - duplicate_penalty


def _detect_header_row(rows: list[list[str]], max_scan_rows: int = 40) -> int:
    best_index = 0
    best_score = 0.0
    for index, row in enumerate(rows[:max_scan_rows]):
        score = _row_score(row)
        next_non_empty = _non_empty_count(rows[index + 1]) if index + 1 < len(rows) else 0
        if next_non_empty >= 2:
            score += min(next_non_empty, 8) * 0.15
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def _merged_lookup(sheet: Any) -> dict[tuple[int, int], str]:
    if not hasattr(sheet, "merged_cells"):
        return {}
    lookup: dict[tuple[int, int], str] = {}
    for merged_range in sheet.merged_cells.ranges:
        top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        text = _cell_to_text(top_left)
        if not text:
            continue
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(row_index, column_index)] = text
    return lookup


def _extract_rows(sheet: Any, max_rows: int = 20000, max_cols: int = 200) -> tuple[list[list[str]], dict[str, Any]]:
    if not hasattr(sheet, "cell"):
        return _extract_rows_streaming(sheet, max_rows=max_rows, max_cols=max_cols)
    merged_values = _merged_lookup(sheet)
    rows: list[list[str]] = []
    scanned_rows = min(sheet.max_row or 0, max_rows)
    scanned_cols = min(sheet.max_column or 0, max_cols)
    first_used_col: int | None = None
    last_used_col = 0

    for row_index in range(1, scanned_rows + 1):
        values = []
        for column_index in range(1, scanned_cols + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            text = _cell_to_text(value) or merged_values.get((row_index, column_index), "")
            values.append(text)
            if text:
                first_used_col = column_index if first_used_col is None else min(first_used_col, column_index)
                last_used_col = max(last_used_col, column_index)
        if _non_empty_count(values):
            rows.append(values)

    if first_used_col is None:
        return [], {"scanned_rows": scanned_rows, "scanned_cols": scanned_cols, "first_used_col": None, "last_used_col": None}

    left = first_used_col - 1
    right = last_used_col
    trimmed_rows = [row[left:right] for row in rows]
    return trimmed_rows, {
        "scanned_rows": scanned_rows,
        "scanned_cols": scanned_cols,
        "first_used_col": first_used_col,
        "last_used_col": last_used_col,
    }


def _extract_rows_streaming(sheet: Any, max_rows: int = 20000, max_cols: int = 200) -> tuple[list[list[str]], dict[str, Any]]:
    rows: list[list[str]] = []
    scanned_rows = min(sheet.max_row or 0, max_rows)
    scanned_cols = min(sheet.max_column or 0, max_cols)
    first_used_col: int | None = None
    last_used_col = 0

    for values_tuple in sheet.iter_rows(min_row=1, max_row=scanned_rows, max_col=scanned_cols, values_only=True):
        values = [_cell_to_text(value) for value in values_tuple]
        for column_index, text in enumerate(values, start=1):
            if text:
                first_used_col = column_index if first_used_col is None else min(first_used_col, column_index)
                last_used_col = max(last_used_col, column_index)
        if _non_empty_count(values):
            rows.append(values)

    if first_used_col is None:
        return [], {"scanned_rows": scanned_rows, "scanned_cols": scanned_cols, "first_used_col": None, "last_used_col": None, "streaming": True}

    left = first_used_col - 1
    right = last_used_col
    return [row[left:right] for row in rows], {
        "scanned_rows": scanned_rows,
        "scanned_cols": scanned_cols,
        "first_used_col": first_used_col,
        "last_used_col": last_used_col,
        "streaming": True,
    }


def _sheet_formula_cells(formula_sheet: Any, limit: int = 20) -> list[str]:
    if formula_sheet is None:
        return []
    formula_cells = []
    for row in formula_sheet.iter_rows():
        for cell in row:
            value = _cell_to_text(cell.value)
            if value.startswith("="):
                formula_cells.append(f"{cell.coordinate}: {value[:120]}")
                if len(formula_cells) >= limit:
                    return formula_cells
    return formula_cells


def _sheet_profile(sheet: Any, formula_sheet: Any) -> dict[str, Any]:
    rows, scan_meta = _extract_rows(sheet, max_rows=5000, max_cols=120)
    warnings = []
    header_index = _detect_header_row(rows) if rows else 0
    header = rows[header_index] if rows else []
    non_empty_columns = len([column for column in zip(*rows) if any(value for value in column)]) if rows else 0
    duplicate_headers = sorted({value for value in header if value and header.count(value) > 1})
    formula_cells = _sheet_formula_cells(formula_sheet)

    if not rows:
        warnings.append("空工作表或未识别到有效数据。")
    merged_ranges = [str(item) for item in list(sheet.merged_cells.ranges)[:20]] if hasattr(sheet, "merged_cells") else []
    if merged_ranges:
        warnings.append("包含合并单元格，清洗时会用左上角值填充。")
    if duplicate_headers:
        warnings.append("存在重复表头，清洗时会自动加后缀。")
    if formula_cells:
        warnings.append("包含公式，清洗使用当前缓存值；如公式未重算，结果可能不是最新。")
    if rows and _non_empty_count(header) < 2:
        warnings.append("表头不清晰，建议人工指定 header_row。")
    if sheet.max_row and len(rows) / max(sheet.max_row, 1) < 0.5:
        warnings.append("空行较多，清洗时会丢弃空行。")

    return {
        "sheet": sheet.title,
        "dimensions": f"{sheet.max_row} rows x {sheet.max_column} columns",
        "non_empty_rows": len(rows),
        "non_empty_columns": non_empty_columns,
        "candidate_header_row": header_index + 1 if rows else None,
        "candidate_headers": header[:60],
        "merged_ranges": merged_ranges,
        "duplicate_headers": duplicate_headers,
        "formula_cells": formula_cells,
        "sample_rows": rows[header_index + 1 : header_index + 6] if rows else [],
        "warnings": warnings,
        "scan": scan_meta,
    }


def profile_excel_file(path: str) -> str:
    """分析 raw 目录中的 Excel 表结构，识别表头、合并单元格、公式和清洗风险。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel profiling requires openpyxl. Run pip install -r requirements.txt.") from exc

    _ensure_dirs()
    target = _raw_path(path)
    if _is_huge_excel(target):
        return json.dumps(
            {
                "file": str(target),
                "file_size_mb": round(target.stat().st_size / 1024 / 1024, 2),
                "large_file_safe_mode": True,
                "too_large_for_interactive_profile": True,
                "recommendation": (
                    "This workbook is over 100MB. Do not profile it inside a frontend chat request. "
                    "Split/export the needed sheets to smaller CSV/XLSX files first, then place those files in raw."
                ),
            },
            ensure_ascii=False,
            indent=2,
        )
    is_large = _is_large_excel(target)
    try:
        workbook = load_workbook(target, read_only=is_large, data_only=True)
        formula_workbook = None if is_large else load_workbook(target, read_only=False, data_only=False)
    except Exception as exc:
        fallback_sheets = read_xlsx_sheets(target, max_rows=5000, max_cols=120, max_sheets=20)
        sheets = []
        for sheet in fallback_sheets:
            rows = sheet["rows"]
            header_index = _detect_header_row(rows) if rows else 0
            header = rows[header_index] if rows else []
            non_empty_columns = len([column for column in zip(*rows) if any(value for value in column)]) if rows else 0
            sheets.append(
                {
                    "sheet": sheet["sheet"],
                    "dimensions": sheet.get("dimension") or f"{len(rows)} rows x {sheet.get('max_column', 0)} columns",
                    "non_empty_rows": len(rows),
                    "non_empty_columns": non_empty_columns,
                    "candidate_header_row": header_index + 1 if rows else None,
                    "candidate_headers": header[:60],
                    "merged_ranges": [],
                    "duplicate_headers": sorted({value for value in header if value and header.count(value) > 1}),
                    "formula_cells": [],
                    "sample_rows": rows[header_index + 1 : header_index + 6] if rows else [],
                    "warnings": [f"openpyxl failed to read workbook styles; used OOXML value fallback: {exc}"],
                    "scan": {"fallback": "ooxml", "scanned_rows": len(rows), "scanned_cols": sheet.get("max_column", 0)},
                }
            )
        return json.dumps(
            {
                "file": str(target),
                "file_size_mb": round(target.stat().st_size / 1024 / 1024, 2),
                "large_file_safe_mode": True,
                "sheet_count": len(sheets),
                "profiled_sheets": sheets,
                "note": "Used OOXML fallback because openpyxl could not parse workbook styles.",
            },
            ensure_ascii=False,
            indent=2,
        )
    sheets = [
        _sheet_profile(sheet, None if formula_workbook is None else formula_workbook.worksheets[index])
        for index, sheet in enumerate(workbook.worksheets[:20])
    ]
    return json.dumps(
        {
            "file": str(target),
            "file_size_mb": round(target.stat().st_size / 1024 / 1024, 2),
            "large_file_safe_mode": is_large,
            "sheet_count": len(workbook.worksheets),
            "profiled_sheets": sheets,
            "note": "Large file profile scans a preview only; use clean_excel_to_csv with sheet_name/header_row/row_limit for controlled extraction." if is_large else "",
        },
        ensure_ascii=False,
        indent=2,
    )


def _clean_rows(rows: list[list[str]], header_index: int) -> tuple[list[str], list[list[str]], dict[str, Any]]:
    if not rows:
        return [], [], {"dropped_blank_rows": 0, "dropped_blank_columns": 0}

    header = rows[header_index]
    data_rows = rows[header_index + 1 :]
    used_columns = [
        index
        for index, value in enumerate(header)
        if value or any(index < len(row) and row[index] for row in data_rows)
    ]
    used_names: dict[str, int] = {}
    headers = [
        _normalize_header(header[column_index], column_position + 1, used_names)
        for column_position, column_index in enumerate(used_columns)
    ]
    cleaned_rows = []
    dropped_blank_rows = 0
    for row in data_rows:
        cleaned = [row[column_index] if column_index < len(row) else "" for column_index in used_columns]
        if not any(cleaned):
            dropped_blank_rows += 1
            continue
        cleaned_rows.append(cleaned)
    return headers, cleaned_rows, {
        "dropped_blank_rows": dropped_blank_rows,
        "dropped_blank_columns": len(header) - len(used_columns),
    }


def _safe_output_name(stem: str, sheet_name: str) -> str:
    name = f"{stem}__{sheet_name}"
    cleaned = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff_-]+", "_", name).strip("_")
    return cleaned[:120] or "cleaned_table"


def clean_excel_to_csv(path: str, sheet_name: str = "", header_row: int = 0, output_name: str = "", row_limit: int = 0) -> str:
    """清洗 raw 目录中的 Excel 表，输出 UTF-8 CSV 到 data/cleaned。header_row 为 1 基序号，0 表示自动识别；大文件默认最多清洗 50000 行。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel cleaning requires openpyxl. Run pip install -r requirements.txt.") from exc

    _ensure_dirs()
    target = _raw_path(path)
    if _is_huge_excel(target):
        return json.dumps(
            {
                "source": str(target),
                "file_size_mb": round(target.stat().st_size / 1024 / 1024, 2),
                "too_large_for_interactive_clean": True,
                "cleaned_dir": str(CLEANED_DIR),
                "results": [],
                "recommendation": (
                    "This workbook is over 100MB and should be converted or split outside the frontend request. "
                    "Recommended: export the required worksheet as CSV, or split by month/SKU/category, then put the smaller files in raw."
                ),
            },
            ensure_ascii=False,
            indent=2,
        )
    is_large = _is_large_excel(target)
    effective_row_limit = row_limit or (DEFAULT_LARGE_CLEAN_ROW_LIMIT if is_large else 20000)
    try:
        workbook = load_workbook(target, read_only=is_large, data_only=True)
    except Exception as exc:
        fallback_sheets = read_xlsx_sheets(target, max_rows=effective_row_limit, max_cols=200, max_sheets=30)
        if sheet_name:
            fallback_sheets = [sheet for sheet in fallback_sheets if sheet["sheet"] == sheet_name]
            if not fallback_sheets:
                raise ValueError(f"Sheet not found: {sheet_name}.") from exc
        elif fallback_sheets:
            fallback_sheets = [max(fallback_sheets, key=lambda item: len(item["rows"]))]
        results = []
        for sheet in fallback_sheets:
            rows = sheet["rows"]
            if not rows:
                results.append({"sheet": sheet["sheet"], "error": "No readable rows found."})
                continue
            detected_header_index = _detect_header_row(rows)
            header_index = header_row - 1 if header_row > 0 else detected_header_index
            headers, cleaned_rows, clean_meta = _clean_rows(rows, header_index)
            output_stem = _safe_output_name(output_name.strip(), sheet["sheet"]) if output_name.strip() else _safe_output_name(target.stem, sheet["sheet"])
            output_path = _safe_under(CLEANED_DIR / f"{output_stem}.csv", [CLEANED_DIR])
            with output_path.open("w", encoding="utf-8-sig", newline="") as file:
                writer = csv.writer(file)
                writer.writerow(headers)
                writer.writerows(cleaned_rows)
            results.append(
                {
                    "sheet": sheet["sheet"],
                    "output_csv": str(output_path),
                    "rows": len(cleaned_rows),
                    "columns": len(headers),
                    "headers": headers,
                    "detected_header_row": detected_header_index + 1,
                    "used_header_row": header_index + 1,
                    "scan": {"fallback": "ooxml", "scanned_rows": len(rows), "scanned_cols": sheet.get("max_column", 0)},
                    "large_file_safe_mode": True,
                    "partial_output": len(rows) >= effective_row_limit,
                    "row_limit": effective_row_limit,
                    "note": f"Used OOXML fallback because openpyxl could not parse workbook styles: {exc}",
                    **clean_meta,
                }
            )
        return json.dumps({"source": str(target), "cleaned_dir": str(CLEANED_DIR), "results": results}, ensure_ascii=False, indent=2)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}. Available: {', '.join(workbook.sheetnames)}")
        sheets = [workbook[sheet_name]]
    else:
        profiles = [_sheet_profile(sheet, None if is_large else sheet) for sheet in workbook.worksheets]
        best_sheet_name = max(profiles, key=lambda item: item["non_empty_rows"])["sheet"]
        sheets = [workbook[best_sheet_name]]

    results = []
    for sheet in sheets:
        rows, scan_meta = _extract_rows(sheet, max_rows=effective_row_limit)
        if not rows:
            results.append({"sheet": sheet.title, "error": "No readable rows found."})
            continue
        detected_header_index = _detect_header_row(rows)
        header_index = header_row - 1 if header_row > 0 else detected_header_index
        if header_index < 0 or header_index >= len(rows):
            raise ValueError(f"Invalid header_row {header_row} for sheet {sheet.title}.")
        headers, cleaned_rows, clean_meta = _clean_rows(rows, header_index)
        output_stem = _safe_output_name(output_name.strip(), sheet.title) if output_name.strip() else _safe_output_name(target.stem, sheet.title)
        if len(sheets) > 1 and output_name.strip():
            output_stem = _safe_output_name(output_stem, sheet.title)
        output_path = _safe_under(CLEANED_DIR / f"{output_stem}.csv", [CLEANED_DIR])
        with output_path.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.writer(file)
            writer.writerow(headers)
            writer.writerows(cleaned_rows)
        results.append(
            {
                "sheet": sheet.title,
                "output_csv": str(output_path),
                "rows": len(cleaned_rows),
                "columns": len(headers),
                "headers": headers,
                "detected_header_row": detected_header_index + 1,
                "used_header_row": header_index + 1,
                "scan": scan_meta,
                "large_file_safe_mode": is_large,
                "partial_output": bool(is_large and (sheet.max_row or 0) > effective_row_limit),
                "row_limit": effective_row_limit,
                "note": "Large file output is limited to row_limit rows to avoid freezing the frontend. Increase row_limit only when you are ready for a long-running extraction." if is_large else "",
                **clean_meta,
            }
        )
    return json.dumps({"source": str(target), "cleaned_dir": str(CLEANED_DIR), "results": results}, ensure_ascii=False, indent=2)


def clean_all_excel_files(limit: int = 20) -> str:
    """批量清洗 raw 目录中的 Excel 文件，每个文件默认选择数据最多的工作表。"""
    _ensure_dirs()
    files = [
        path
        for path in sorted(RAW_DIR.rglob("*"))
        if path.is_file() and path.suffix.lower() in EXCEL_SUFFIXES
    ][:limit]
    results = []
    for path in files:
        try:
            if _is_large_excel(path):
                results.append(
                    {
                        "source": str(path),
                        "file_size_mb": round(path.stat().st_size / 1024 / 1024, 2),
                        "skipped_for_large_file_pipeline": True,
                        "recommendation": "Use large_excel_pipeline for this workbook; batch cleaning only handles small interactive files.",
                    }
                )
                continue
            results.append(json.loads(clean_excel_to_csv(str(path.relative_to(RAW_DIR)))))
        except Exception as exc:
            results.append({"source": str(path), "error": str(exc)})
    return json.dumps({"processed": len(results), "results": results}, ensure_ascii=False, indent=2)


def write_cleaning_report(title: str, content: str) -> str:
    """把表格画像、清洗结果或人工规则保存到 Obsidian logs。"""
    _ensure_dirs()
    slug = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff_-]+", "-", title).strip("-")[:70] or "table-cleaning"
    path = _safe_under(WIKI_DIR / "logs" / f"{datetime.now().strftime('%Y%m%d-%H%M%S')}-{slug}.md", [WIKI_DIR])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(f"# {title}\n\n{content.strip()}\n", encoding="utf-8")
    return json.dumps({"saved_to": str(path)}, ensure_ascii=False, indent=2)
