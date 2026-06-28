from __future__ import annotations

import csv
import json
import os
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from src.a2a_ecommerce_demo.fact_layer_tools import register_large_excel_dataset
from src.a2a_ecommerce_demo.state_io import atomic_write_json

PROJECT_ROOT = Path(__file__).resolve().parents[2]
RAW_DIR = Path(os.getenv("A2A_RAW_DIR", PROJECT_ROOT / "raw")).resolve()
DATA_DIR = Path(os.getenv("A2A_DATA_DIR", PROJECT_ROOT / "data")).resolve()
STAGING_DIR = Path(os.getenv("A2A_STAGING_DIR", DATA_DIR / "staging")).resolve()
WAREHOUSE_DIR = Path(os.getenv("A2A_WAREHOUSE_DIR", DATA_DIR / "warehouse")).resolve()
WIKI_DIR = Path(os.getenv("A2A_WIKI_DIR", PROJECT_ROOT / "wiki")).resolve()

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
LARGE_EXCEL_BYTES = 50 * 1024 * 1024
HUGE_EXCEL_BYTES = 100 * 1024 * 1024

FIELD_GROUP_ALIASES = {
    "sku": ["sku", "msku", "asin", "货品编码", "商品编码", "货品id", "货号", "条码"],
    "product_name": ["product_name", "商品名称", "产品名称", "货品名称", "name", "title"],
    "date": ["date", "日期", "日期范围", "月份", "账期"],
    "warehouse": ["warehouse", "warehouse_code", "仓库", "仓库名称", "仓库code"],
    "opening_inventory": ["期初总量", "期初库存", "opening_inventory", "begin_stock"],
    "inbound": ["入库总量", "入库", "采购入库", "inbound"],
    "outbound": ["出库总量", "出库", "销售发货", "良品销售发货", "outbound", "sales_qty"],
    "ending_inventory": ["期末总量", "期末库存", "库存", "ending_inventory", "stock"],
    "in_transit": ["在途", "采购在途", "调拨在途", "in_transit"],
}


def _ensure_dirs() -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    STAGING_DIR.mkdir(parents=True, exist_ok=True)
    WAREHOUSE_DIR.mkdir(parents=True, exist_ok=True)
    WIKI_DIR.mkdir(parents=True, exist_ok=True)
    (WIKI_DIR / "logs").mkdir(parents=True, exist_ok=True)
    (WIKI_DIR / "data-dictionary").mkdir(parents=True, exist_ok=True)


def _safe_under(path: Path, roots: list[Path]) -> Path:
    resolved = path.resolve()
    if not any(root in [resolved, *resolved.parents] for root in roots):
        raise ValueError(f"Refusing to access outside allowed directories: {resolved}")
    return resolved


def _raw_path(path: str) -> Path:
    target = Path(path)
    if not target.is_absolute():
        if target.parts and target.parts[0].lower() == "raw":
            target = Path(*target.parts[1:])
        target = RAW_DIR / target
    target = _safe_under(target, [RAW_DIR])
    if target.suffix.lower() not in EXCEL_SUFFIXES:
        raise ValueError(f"Only .xlsx/.xlsm files are supported: {target.name}")
    if not target.exists():
        raise FileNotFoundError(f"Raw Excel file not found: {target}")
    return target


def _slugify(value: str, fallback: str = "large_excel") -> str:
    slug = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff_-]+", "_", value).strip("_")
    return slug[:90] or fallback


def _relative(path: Path) -> str:
    for root_name, root in [
        ("raw", RAW_DIR),
        ("warehouse", WAREHOUSE_DIR),
        ("staging", STAGING_DIR),
        ("wiki", WIKI_DIR),
        ("data", DATA_DIR),
    ]:
        try:
            return f"{root_name}/{path.relative_to(root).as_posix()}"
        except ValueError:
            continue
    return path.as_posix()


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _non_empty_count(values: list[str]) -> int:
    return sum(1 for value in values if value)


def _looks_numeric(value: str) -> bool:
    if not value:
        return False
    try:
        float(value.replace(",", ""))
        return True
    except ValueError:
        return False


def _row_score(values: list[str]) -> float:
    non_empty = _non_empty_count(values)
    if non_empty < 2:
        return 0.0
    text_count = sum(1 for value in values if value and not _looks_numeric(value))
    unique_count = len({value for value in values if value})
    return non_empty + text_count * 0.7 - max(0, non_empty - unique_count) * 0.2


def _detect_header_row(rows: list[list[str]], max_scan_rows: int = 50) -> int:
    best_index = 0
    best_score = 0.0
    for index, row in enumerate(rows[:max_scan_rows]):
        score = _row_score(row)
        if index + 1 < len(rows):
            score += min(_non_empty_count(rows[index + 1]), 8) * 0.15
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def _normalize_header(value: str, fallback_index: int, used: dict[str, int]) -> str:
    cleaned = re.sub(r"\s+", "_", value.strip())
    cleaned = re.sub(r"[^\w\u4e00-\u9fff]+", "_", cleaned, flags=re.UNICODE).strip("_")
    cleaned = cleaned[:70] or f"column_{fallback_index}"
    count = used.get(cleaned, 0)
    used[cleaned] = count + 1
    return f"{cleaned}_{count + 1}" if count else cleaned


def _sample_sheet_rows(sheet: Any, sample_rows: int, max_cols: int) -> list[list[str]]:
    rows: list[list[str]] = []
    for values_tuple in sheet.iter_rows(max_row=sample_rows, max_col=max_cols, values_only=True):
        values = [_cell_to_text(value) for value in values_tuple]
        if _non_empty_count(values):
            rows.append(values)
    return rows


def _used_columns(sample_rows: list[list[str]], header_index: int) -> list[int]:
    if not sample_rows:
        return []
    width = max(len(row) for row in sample_rows)
    header = sample_rows[header_index] if header_index < len(sample_rows) else sample_rows[0]
    used = []
    for index in range(width):
        has_header = index < len(header) and bool(header[index])
        has_data = any(index < len(row) and row[index] for row in sample_rows[header_index + 1 :])
        if has_header or has_data:
            used.append(index)
    return used


def _find_field(headers: Sequence[str], group: str) -> str:
    aliases = [alias.lower() for alias in FIELD_GROUP_ALIASES.get(group, [])]
    lowered_headers = [(header, header.lower()) for header in headers]
    for alias in aliases:
        for header, lowered in lowered_headers:
            if alias == lowered:
                return header
    for alias in sorted(aliases, key=len, reverse=True):
        for header, lowered in lowered_headers:
            if alias in lowered:
                return header
    return ""


def _as_float(value: str) -> float | None:
    if value is None:
        return None
    text = str(value).replace(",", "").replace("%", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _write_json(path: Path, data: dict[str, Any]) -> None:
    atomic_write_json(path, data)


def _write_wiki_page(path: Path, title: str, lines: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    body = "\n".join([f"# {title}", "", *lines]).rstrip() + "\n"
    path.write_text(body, encoding="utf-8")


def profile_large_excel_file(path: str, sample_rows_per_sheet: int = 80, max_sheets: int = 30, max_cols: int = 160) -> str:
    """离线体检大 Excel：读取 sheet、尺寸、候选表头和样例行，并保存 profile JSON/wiki 报告。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Large Excel profiling requires openpyxl. Run pip install -r requirements.txt.") from exc

    _ensure_dirs()
    target = _raw_path(path)
    workbook = load_workbook(target, read_only=True, data_only=True)
    file_slug = _slugify(target.stem)
    profile_dir = _safe_under(STAGING_DIR / "large_excel_profiles", [STAGING_DIR])
    sheets = []
    try:
        for sheet in workbook.worksheets[:max_sheets]:
            sample_rows = _sample_sheet_rows(sheet, sample_rows_per_sheet, max_cols)
            header_index = _detect_header_row(sample_rows) if sample_rows else 0
            header = sample_rows[header_index] if sample_rows else []
            sheets.append(
                {
                    "sheet": sheet.title,
                    "dimensions": f"{sheet.max_row or 0} rows x {sheet.max_column or 0} columns",
                    "candidate_header_row": header_index + 1 if sample_rows else None,
                    "candidate_headers": header[:80],
                    "sample_rows": sample_rows[header_index + 1 : header_index + 6],
                    "warnings": [
                        "超过 100MB，已进入离线大文件模式。" if target.stat().st_size >= HUGE_EXCEL_BYTES else "超过 50MB，建议离线处理。",
                        "只抽样读取前几行用于体检，完整数据由后台分块转换。",
                    ],
                }
            )
    finally:
        workbook.close()

    profile = {
        "schema": "a2a_large_excel_profile_v1",
        "file": str(target),
        "relative_path": _relative(target),
        "file_size_mb": round(target.stat().st_size / 1024 / 1024, 2),
        "profiled_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sheet_count": len(workbook.worksheets),
        "profiled_sheets": sheets,
        "recommendation": "Use process_large_excel_file to build the dataset registry, Parquet fact layer, and dataset wiki pages before decision analysis.",
    }
    profile_path = profile_dir / f"{file_slug}__profile.json"
    _write_json(profile_path, profile)

    wiki_path = WIKI_DIR / "logs" / f"Large-Excel-Profile-{file_slug}.md"
    lines = [
        f"- Source: `{_relative(target)}`",
        f"- Size: {profile['file_size_mb']} MB",
        f"- Sheets: {profile['sheet_count']}",
        f"- Profile JSON: `{_relative(profile_path)}`",
        "",
        "## Sheets",
    ]
    for sheet in sheets:
        lines.extend(
            [
                "",
                f"### {sheet['sheet']}",
                f"- Dimensions: {sheet['dimensions']}",
                f"- Candidate header row: {sheet['candidate_header_row']}",
                f"- Candidate headers: {', '.join(str(item) for item in sheet['candidate_headers'][:30])}",
            ]
        )
    _write_wiki_page(wiki_path, f"Large Excel Profile - {target.stem}", lines)
    profile["profile_path"] = str(profile_path)
    profile["wiki_path"] = _relative(wiki_path)
    return json.dumps(profile, ensure_ascii=False, indent=2)


def process_large_excel_file(
    path: str,
    rows_per_chunk: int = 50000,
    max_sheets: int = 30,
    max_cols: int = 220,
    sample_rows_per_sheet: int = 200,
) -> str:
    """把大 Excel 离线拆分为可追溯 CSV chunk，并生成 manifest 与 Obsidian 摘要。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Large Excel processing requires openpyxl. Run pip install -r requirements.txt.") from exc

    _ensure_dirs()
    target = _raw_path(path)
    file_slug = _slugify(target.stem)
    output_dir = _safe_under(WAREHOUSE_DIR / "large_excel" / file_slug, [WAREHOUSE_DIR])
    output_dir.mkdir(parents=True, exist_ok=True)
    for stale in output_dir.glob("*.csv"):
        stale.unlink()

    workbook = load_workbook(target, read_only=True, data_only=True)
    manifest: dict[str, Any] = {
        "schema": "a2a_large_excel_manifest_v1",
        "source": str(target),
        "relative_source": _relative(target),
        "file_size_mb": round(target.stat().st_size / 1024 / 1024, 2),
        "processed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "rows_per_chunk": rows_per_chunk,
        "output_dir": str(output_dir),
        "sheets": [],
        "chunks": [],
    }

    try:
        for sheet in workbook.worksheets[:max_sheets]:
            sample_rows = _sample_sheet_rows(sheet, sample_rows_per_sheet, max_cols)
            if not sample_rows:
                manifest["sheets"].append({"sheet": sheet.title, "status": "skipped", "reason": "no readable rows"})
                continue
            header_index = _detect_header_row(sample_rows)
            used_columns = _used_columns(sample_rows, header_index)
            if not used_columns:
                manifest["sheets"].append({"sheet": sheet.title, "status": "skipped", "reason": "no usable columns"})
                continue
            header_row = sample_rows[header_index]
            used_names: dict[str, int] = {}
            headers = [
                _normalize_header(header_row[index] if index < len(header_row) else "", position + 1, used_names)
                for position, index in enumerate(used_columns)
            ]
            headers_with_trace = ["_source_file", "_source_sheet", "_source_row_number", *headers]
            sheet_slug = _slugify(sheet.title, "sheet")
            chunk_index = 1
            current_rows = 0
            total_rows = 0
            writer = None
            file_handle = None
            chunk_path: Path | None = None

            def open_chunk() -> None:
                nonlocal writer, file_handle, chunk_path, current_rows, chunk_index
                if file_handle:
                    file_handle.close()
                chunk_path = output_dir / f"{file_slug}__{sheet_slug}__part_{chunk_index:04d}.csv"
                file_handle = chunk_path.open("w", encoding="utf-8-sig", newline="")
                writer = csv.writer(file_handle)
                writer.writerow(headers_with_trace)
                current_rows = 0

            open_chunk()
            for row_number, values_tuple in enumerate(sheet.iter_rows(max_col=max_cols, values_only=True), start=1):
                if row_number <= header_index + 1:
                    continue
                values = [_cell_to_text(value) for value in values_tuple]
                cleaned = [values[index] if index < len(values) else "" for index in used_columns]
                if not any(cleaned):
                    continue
                if current_rows >= rows_per_chunk:
                    if file_handle:
                        file_handle.close()
                    manifest["chunks"].append(
                        {
                            "sheet": sheet.title,
                            "path": _relative(chunk_path) if chunk_path else "",
                            "rows": current_rows,
                            "columns": len(headers),
                        }
                    )
                    chunk_index += 1
                    open_chunk()
                assert writer is not None
                writer.writerow([target.name, sheet.title, row_number, *cleaned])
                current_rows += 1
                total_rows += 1

            if file_handle:
                file_handle.close()
            if chunk_path and current_rows:
                manifest["chunks"].append(
                    {
                        "sheet": sheet.title,
                        "path": _relative(chunk_path),
                        "rows": current_rows,
                        "columns": len(headers),
                    }
                )
            elif chunk_path and chunk_path.exists():
                chunk_path.unlink()
            manifest["sheets"].append(
                {
                    "sheet": sheet.title,
                    "status": "processed",
                    "detected_header_row": header_index + 1,
                    "rows": total_rows,
                    "columns": len(headers),
                    "headers": headers,
                    "chunks": chunk_index if total_rows else 0,
                }
            )
    finally:
        workbook.close()

    manifest_path = output_dir / "manifest.json"
    _write_json(manifest_path, manifest)
    quality = json.loads(assess_large_excel_quality(str(manifest_path)))
    wiki_path = WIKI_DIR / "logs" / f"Large-Excel-Processing-{file_slug}.md"
    lines = [
        f"- Source: `{_relative(target)}`",
        f"- Manifest: `{_relative(manifest_path)}`",
        f"- Output dir: `{_relative(output_dir)}`",
        f"- Chunks: {len(manifest['chunks'])}",
        f"- Rows: {sum(chunk.get('rows', 0) for chunk in manifest['chunks'])}",
        f"- Quality level: {quality.get('quality_level')}",
        "",
        "## Sheets",
    ]
    for sheet in manifest["sheets"]:
        lines.append(f"- {sheet.get('sheet')}: {sheet.get('status')} / rows={sheet.get('rows', 0)} / columns={sheet.get('columns', 0)}")
    lines.extend(["", "## Quality Findings"])
    for warning in quality.get("warnings", [])[:20]:
        lines.append(f"- {warning}")
    _write_wiki_page(wiki_path, f"Large Excel Processing - {target.stem}", lines)
    manifest["manifest_path"] = str(manifest_path)
    manifest["wiki_path"] = _relative(wiki_path)
    manifest["quality"] = quality
    try:
        dataset_registry = register_large_excel_dataset(str(manifest_path), quality.get("quality_report_path", ""))
    except Exception as exc:
        dataset_registry = {
            "status": "warning",
            "reason": f"Failed to register DuckDB fact layer: {exc}",
        }
    manifest["dataset_registry"] = dataset_registry
    return json.dumps(manifest, ensure_ascii=False, indent=2)


def assess_large_excel_quality(manifest_path: str) -> str:
    """检查大 Excel chunk 的字段完整性、空值、重复键和库存平衡。"""
    _ensure_dirs()
    target = Path(manifest_path)
    if not target.is_absolute():
        target = WAREHOUSE_DIR / "large_excel" / manifest_path
    target = _safe_under(target, [WAREHOUSE_DIR, STAGING_DIR, DATA_DIR])
    manifest = json.loads(target.read_text(encoding="utf-8"))
    total_rows = 0
    field_counter: Counter[str] = Counter()
    empty_counter: Counter[str] = Counter()
    duplicate_key_count = 0
    seen_keys: set[str] = set()
    balance_checked = 0
    balance_failed = 0
    warnings: list[str] = []
    matched_groups: set[str] = set()

    for chunk in manifest.get("chunks", []):
        chunk_path = Path(str(chunk.get("path", "")))
        if not chunk_path.is_absolute():
            if chunk_path.parts and chunk_path.parts[0].lower() == "warehouse":
                chunk_path = WAREHOUSE_DIR / Path(*chunk_path.parts[1:])
            elif chunk_path.parts and chunk_path.parts[0].lower() == "data":
                chunk_path = DATA_DIR / Path(*chunk_path.parts[1:])
            else:
                chunk_path = DATA_DIR / chunk_path
        chunk_path = _safe_under(chunk_path, [WAREHOUSE_DIR, DATA_DIR])
        with chunk_path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file)
            headers = reader.fieldnames or []
            for group in FIELD_GROUP_ALIASES:
                if _find_field(headers, group):
                    matched_groups.add(group)
            opening_field = _find_field(headers, "opening_inventory")
            inbound_field = _find_field(headers, "inbound")
            outbound_field = _find_field(headers, "outbound")
            ending_field = _find_field(headers, "ending_inventory")
            sku_field = _find_field(headers, "sku")
            date_field = _find_field(headers, "date")
            warehouse_field = _find_field(headers, "warehouse")
            for row in reader:
                total_rows += 1
                for header in headers:
                    field_counter[header] += 1
                    if not str(row.get(header, "")).strip():
                        empty_counter[header] += 1
                if sku_field or date_field or warehouse_field:
                    key = "|".join(str(row.get(field, "")).strip() for field in [sku_field, date_field, warehouse_field] if field)
                    if key and key in seen_keys:
                        duplicate_key_count += 1
                    if key:
                        seen_keys.add(key)
                if opening_field and inbound_field and outbound_field and ending_field:
                    opening = _as_float(row.get(opening_field, ""))
                    inbound = _as_float(row.get(inbound_field, "")) or 0.0
                    outbound = _as_float(row.get(outbound_field, "")) or 0.0
                    ending = _as_float(row.get(ending_field, ""))
                    if opening is not None and ending is not None:
                        balance_checked += 1
                        expected_add = opening + inbound + outbound
                        expected_subtract = opening + inbound - outbound
                        if min(abs(expected_add - ending), abs(expected_subtract - ending)) > 1e-6:
                            balance_failed += 1

    high_empty_fields = [
        {"field": field, "empty_rate": round(empty_counter[field] / count, 4)}
        for field, count in field_counter.items()
        if count and empty_counter[field] / count >= 0.5 and not field.startswith("_source_")
    ][:30]
    if high_empty_fields:
        warnings.append(f"存在高空值字段：{', '.join(item['field'] for item in high_empty_fields[:10])}。")
    if duplicate_key_count:
        warnings.append(f"发现 {duplicate_key_count} 行可能重复的 SKU+日期+仓库组合。")
    if balance_checked and balance_failed:
        warnings.append(f"库存平衡校验异常：{balance_failed}/{balance_checked} 行不满足 期初+入库±出库≈期末。")
    required = {"sku", "product_name", "date", "warehouse", "ending_inventory", "outbound"}
    missing_groups = sorted(required - matched_groups)
    if missing_groups:
        warnings.append(f"缺少库存决策常用字段组：{', '.join(missing_groups)}。")
    quality_level = "high"
    if warnings:
        quality_level = "medium"
    if len(missing_groups) >= 3 or total_rows == 0:
        quality_level = "low"

    result = {
        "schema": "a2a_large_excel_quality_v1",
        "manifest_path": str(target),
        "total_rows": total_rows,
        "chunks": len(manifest.get("chunks", [])),
        "matched_field_groups": sorted(matched_groups),
        "missing_field_groups": missing_groups,
        "high_empty_fields": high_empty_fields,
        "duplicate_sku_date_warehouse_rows": duplicate_key_count,
        "inventory_balance": {
            "checked_rows": balance_checked,
            "failed_rows": balance_failed,
            "failure_rate": round(balance_failed / balance_checked, 4) if balance_checked else 0,
        },
        "warnings": warnings,
        "quality_level": quality_level,
    }
    quality_path = target.parent / "quality_report.json"
    _write_json(quality_path, result)
    result["quality_report_path"] = str(quality_path)
    return json.dumps(result, ensure_ascii=False, indent=2)


def process_all_large_excel_files(limit: int = 10, min_size_mb: int = 50, rows_per_chunk: int = 50000) -> str:
    """批量处理 raw 中的大 Excel，输出 profile、chunk、quality 和 wiki 摘要。"""
    _ensure_dirs()
    min_bytes = min_size_mb * 1024 * 1024
    files = [
        path
        for path in sorted(RAW_DIR.rglob("*"))
        if path.is_file() and path.suffix.lower() in EXCEL_SUFFIXES and path.stat().st_size >= min_bytes
    ][:limit]
    results = []
    for path in files:
        try:
            profile = json.loads(profile_large_excel_file(_relative(path)))
            processed = json.loads(process_large_excel_file(_relative(path), rows_per_chunk=rows_per_chunk))
            results.append({"file": _relative(path), "status": "success", "profile": profile, "processed": processed})
        except Exception as exc:
            results.append({"file": _relative(path), "status": "failed", "error": str(exc)})
    return json.dumps(
        {
            "status": "success" if not any(item.get("status") == "failed" for item in results) else "warning",
            "min_size_mb": min_size_mb,
            "processed": len(results),
            "results": results,
        },
        ensure_ascii=False,
        indent=2,
    )
