from __future__ import annotations

import csv
import json
import math
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from src.a2a_ecommerce_demo.fact_layer_tools import (
    audit_fact_source_readiness as audit_fact_source_readiness_from_registry,
)
from src.a2a_ecommerce_demo.fact_layer_tools import (
    duckdb_installed,
)
from src.a2a_ecommerce_demo.fact_layer_tools import (
    list_fact_tables as list_fact_tables_from_registry,
)
from src.a2a_ecommerce_demo.fact_layer_tools import (
    list_registered_datasets as list_registered_datasets_from_registry,
)
from src.a2a_ecommerce_demo.fact_layer_tools import (
    query_fact_layer as query_fact_layer_from_registry,
)
from src.a2a_ecommerce_demo.fact_layer_tools import (
    register_all_fact_datasets as register_all_fact_datasets_from_registry,
)

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = Path(os.getenv("A2A_DATA_DIR", PROJECT_ROOT / "data")).resolve()
WAREHOUSE_DIR = Path(os.getenv("A2A_WAREHOUSE_DIR", DATA_DIR / "warehouse")).resolve()
REPORT_DIR = DATA_DIR / "reports"
MAX_ROWS_PER_DATA_FILE = int(os.getenv("A2A_MAX_ROWS_PER_DATA_FILE", "20000"))
MAX_WAREHOUSE_ROWS_PER_FILE = int(os.getenv("A2A_MAX_WAREHOUSE_ROWS_PER_FILE", "5000"))
_SKIPPED_DATA_DIRS = {"audit", "backups", "derived", "lightrag", "staging", "tasks"}


DATASET_ALIASES = {
    "inventory": ["inventory", "stock", "库存", "进销存", "台账"],
    "sales": ["sales", "销量", "销售", "orders", "订单"],
    "ads": ["ads", "advertising", "广告", "acos", "roas"],
    "profit": ["profit", "finance", "财务", "利润", "margin", "现金流", "损益", "收入", "成本"],
    "products": ["products", "catalog", "商品", "产品", "sku"],
    "suppliers": ["supplier", "suppliers", "供应商", "采购", "工厂"],
}

FIELD_GROUP_ALIASES = {
    "sku": ["sku", "msku", "seller_sku", "asin", "商品编码", "货品编码", "货品编号", "货号", "货品id"],
    "product_name": ["product_name", "name", "title", "商品名称", "产品名称", "货品名称"],
    "date": ["date", "日期", "日期范围", "order_date", "账期", "月份", "发货时间", "建单时间", "货品级发货时间"],
    "inventory": ["current_stock", "stock", "available", "可售库存", "期末库存", "期末总量", "期末良品", "库存", "可用库存", "库存数量", "求和项_库存数量", "求和项_可用库存"],
    "inbound": ["inbound_stock", "inbound", "on_the_way", "在途", "在途库存", "采购在途", "调拨在途", "入库申请"],
    "sales_qty": ["sales_qty", "units_sold", "销量", "销售数量", "出库总量", "良品销售发货", "出库", "数量", "近30天销量"],
    "revenue": ["revenue", "sales_amount", "gmv", "销售额", "收入", "营业收入", "销售收入", "分摊后金额"],
    "cost": ["cost", "cogs", "采购成本", "货品成本", "主营业务成本", "成本"],
    "gross_profit": ["gross_profit", "profit", "毛利", "利润", "gross_margin_profit"],
    "cash": ["cash", "cash_balance", "现金", "现金余额", "可用资金", "账户余额"],
    "ad_spend": ["ad_spend", "spend", "广告花费", "广告支出", "推广费"],
    "acos": ["acos", "acos_7d", "广告销售成本比"],
    "roas": ["roas", "roas_7d", "广告投入产出比"],
    "supplier": ["supplier", "供应商", "工厂", "vendor"],
    "warehouse": ["warehouse", "warehouse_code", "仓库", "仓库名称", "仓库code", "发货仓库"],
    "lead_time": ["lead_time_days", "purchase_lead_time", "补货周期", "采购周期", "交期"],
}


def _sql_string(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def _ensure_data_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)


def _fact_tables() -> set[tuple[str, str]]:
    try:
        result = json.loads(list_fact_tables_from_registry())
    except Exception:
        return set()
    if not result.get("available", False):
        return set()
    return {(str(item.get("schema", "")), str(item.get("name", ""))) for item in result.get("tables", [])}


def _has_fact_table(name: str, schema: str = "marts") -> bool:
    return (schema, name) in _fact_tables()


def _query_fact_rows(sql: str, limit: int = 200) -> list[dict[str, Any]]:
    try:
        result = json.loads(query_fact_layer_from_registry(sql, limit=limit))
    except Exception:
        return []
    return list(result.get("rows", []))


def _fact_layer_summary() -> dict[str, Any]:
    try:
        datasets = json.loads(list_registered_datasets_from_registry()) if duckdb_installed() else {"datasets": []}
    except Exception:
        datasets = {"datasets": [], "registry_path": str(_dataset_registry_path())}
    try:
        tables = json.loads(list_fact_tables_from_registry()) if duckdb_installed() else {"available": False, "tables": []}
    except Exception as exc:
        tables = {"available": False, "tables": [], "reason": str(exc)}
    return {
        "available": bool(tables.get("available")),
        "registry_path": datasets.get("registry_path", ""),
        "duckdb_path": tables.get("duckdb_path", datasets.get("duckdb_path", "")),
        "datasets": datasets.get("datasets", []),
        "tables": tables.get("tables", []),
    }


def _dataset_registry_path() -> Path:
    return Path(os.getenv("A2A_DATASET_REGISTRY", WAREHOUSE_DIR / "dataset_registry.json")).resolve()


def _load_dataset_registry() -> dict[str, Any]:
    registry_path = _dataset_registry_path()
    if not registry_path.exists():
        return {"datasets": {}}
    try:
        return json.loads(registry_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"datasets": {}}


def _unique_non_empty(items: list[str]) -> list[str]:
    seen = set()
    unique = []
    for item in items:
        text = str(item or "").strip()
        if text and text not in seen:
            seen.add(text)
            unique.append(text)
    return unique


def _decision_evidence_chain(
    quality: dict[str, Any] | None = None,
    report_path: str = "",
    extra_evidence: list[str] | None = None,
) -> dict[str, Any]:
    quality = quality or {}
    extra_evidence = extra_evidence or []
    fact_layer = quality.get("fact_layer") or _fact_layer_summary()
    registry = _load_dataset_registry()
    registry_datasets = list(registry.get("datasets", {}).values())

    wiki_pages: list[str] = []
    manifest_paths: list[str] = []
    report_paths: list[str] = []
    mart_fields: dict[str, set[str]] = {}

    for dataset in registry_datasets:
        wiki_pages.extend(str(path) for path in dataset.get("wiki_pages", {}).values())
        if dataset.get("manifest_path"):
            manifest_paths.append(str(dataset["manifest_path"]))
        if dataset.get("quality_report_path"):
            report_paths.append(str(dataset["quality_report_path"]))
        fields = {
            str(header)
            for sheet in dataset.get("sheet_views", [])
            for header in sheet.get("headers", [])
            if str(header).strip()
        }
        for mart in dataset.get("mart_views", []) + dataset.get("semantic_views", []):
            mart_name = str(mart.get("category") or mart.get("view_name") or "").strip()
            if mart_name:
                mart_fields.setdefault(mart_name, set()).update(fields)

    for dataset in fact_layer.get("datasets", []):
        if dataset.get("overview_page"):
            wiki_pages.append(str(dataset["overview_page"]))
        for mart in dataset.get("mart_views", []):
            mart_name = str(mart).strip()
            if mart_name:
                mart_fields.setdefault(mart_name, set())

    for table in fact_layer.get("tables", []):
        if table.get("schema") == "marts" and table.get("name"):
            mart_fields.setdefault(str(table["name"]), set())

    for item in extra_evidence:
        text = str(item or "").strip()
        if not text:
            continue
        if text.startswith("wiki/") or "/wiki/" in text:
            wiki_pages.append(text)
        if text.endswith("manifest.json"):
            manifest_paths.append(text)
        if "quality_report" in text or "report" in Path(text).name.lower():
            report_paths.append(text)

    if report_path:
        report_paths.append(report_path)

    return {
        "wiki_pages": _unique_non_empty(wiki_pages),
        "duckdb_path": str(fact_layer.get("duckdb_path") or os.getenv("A2A_DUCKDB_PATH", WAREHOUSE_DIR / "a2a.duckdb")),
        "registry_path": str(fact_layer.get("registry_path") or _dataset_registry_path()),
        "duckdb_marts": [
            {"mart": mart, "fields": sorted(fields)[:80]}
            for mart, fields in sorted(mart_fields.items())
        ],
        "data_gaps": _unique_non_empty([str(item) for item in quality.get("missing_field_groups", [])]),
        "manifest_paths": _unique_non_empty(manifest_paths),
        "report_paths": _unique_non_empty(report_paths),
    }


def _format_evidence_chain_markdown(evidence_chain: dict[str, Any]) -> str:
    lines = [
        "## Evidence Chain",
        "",
        "### Wiki pages",
    ]
    wiki_pages = evidence_chain.get("wiki_pages", [])
    lines.extend(f"- `{path}`" for path in wiki_pages[:30])
    if not wiki_pages:
        lines.append("- None")

    lines.extend(["", "### DuckDB marts and fields"])
    marts = evidence_chain.get("duckdb_marts", [])
    for mart in marts[:30]:
        fields = ", ".join(mart.get("fields", [])[:20]) or "fields unavailable"
        lines.append(f"- `{mart.get('mart', '')}`: {fields}")
    if not marts:
        lines.append("- None")
    if evidence_chain.get("duckdb_path"):
        lines.append(f"- DuckDB: `{evidence_chain['duckdb_path']}`")
    if evidence_chain.get("registry_path"):
        lines.append(f"- Registry: `{evidence_chain['registry_path']}`")

    lines.extend(["", "### Data gaps"])
    data_gaps = evidence_chain.get("data_gaps", [])
    lines.append(f"- {', '.join(data_gaps) if data_gaps else 'None'}")

    lines.extend(["", "### Manifest/report paths"])
    paths = _unique_non_empty(evidence_chain.get("manifest_paths", []) + evidence_chain.get("report_paths", []))
    lines.extend(f"- `{path}`" for path in paths[:40])
    if not paths:
        lines.append("- None")
    return "\n".join(lines)


def _safe_path(path: Path) -> Path:
    resolved = path.resolve()
    if DATA_DIR not in [resolved, *resolved.parents]:
        raise ValueError(f"Refusing to read outside data directory: {resolved}")
    return resolved


def _is_warehouse_path(path: Path) -> bool:
    resolved = path.resolve()
    return WAREHOUSE_DIR in [resolved, *resolved.parents]


def _should_include_business_file(path: Path) -> bool:
    if not path.is_file() or path.suffix.lower() not in {".csv", ".xlsx", ".xlsm"}:
        return False
    try:
        relative_parts = path.relative_to(DATA_DIR).parts
    except ValueError:
        return False
    if relative_parts and relative_parts[0].lower() in _SKIPPED_DATA_DIRS:
        return False
    return True


def _row_limit_for(path: Path) -> int:
    return MAX_WAREHOUSE_ROWS_PER_FILE if _is_warehouse_path(path) else MAX_ROWS_PER_DATA_FILE


def _normalize_key(key: Any) -> str:
    return str(key or "").strip().lower().replace(" ", "_")


def _should_keep_text(key: Any) -> bool:
    normalized = _normalize_key(key)
    text_markers = [
        "id",
        "sku",
        "asin",
        "fnsku",
        "msku",
        "code",
        "barcode",
        "ean",
        "upc",
        "编码",
        "货号",
        "货品id",
        "条码",
    ]
    return any(marker in normalized for marker in text_markers)


def _coerce_value(value: Any, key: Any = "") -> Any:
    if value is None:
        return ""
    if _should_keep_text(key):
        return str(value).strip()
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return ""
    cleaned = text.replace(",", "").replace("%", "")
    try:
        number = float(cleaned)
        if math.isfinite(number):
            return number
    except ValueError:
        pass
    return text


def _infer_dataset_name(path: Path) -> str:
    lowered = path.stem.lower()
    for dataset, aliases in DATASET_ALIASES.items():
        if any(alias.lower() in lowered for alias in aliases):
            return dataset
    return path.stem.lower()


def _read_csv(path: Path, max_rows: int | None = None) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        rows = []
        for index, row in enumerate(reader):
            if max_rows is not None and index >= max_rows:
                break
            rows.append({_normalize_key(key): _coerce_value(value, key) for key, value in row.items()})
        return rows


def _read_xlsx(path: Path, max_rows: int | None = None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Reading Excel files requires openpyxl. Run pip install -r requirements.txt.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        return []
    row_iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(row_iterator)
    except StopIteration:
        return []
    headers = [_normalize_key(cell) for cell in header_row]
    records: list[dict[str, Any]] = []
    for index, row in enumerate(row_iterator):
        if max_rows is not None and index >= max_rows:
            break
        record = {
            headers[index]: _coerce_value(row[index] if index < len(row) else "", headers[index])
            for index in range(len(headers))
            if headers[index]
        }
        if any(value != "" for value in record.values()):
            records.append(record)
    return records


def _load_table(path: Path, max_rows: int | None = None) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        return _read_csv(path, max_rows=max_rows)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx(path, max_rows=max_rows)
    raise ValueError(f"Unsupported data file: {path.name}")


def _load_business_data() -> dict[str, list[dict[str, Any]]]:
    _ensure_data_dir()
    data: dict[str, list[dict[str, Any]]] = {}
    for path in sorted(DATA_DIR.rglob("*")):
        if not _should_include_business_file(path):
            continue
        safe = _safe_path(path)
        dataset = _infer_dataset_name(safe)
        data.setdefault(dataset, []).extend(_load_table(safe, max_rows=_row_limit_for(safe)))
    return data


def _find_value(record: dict[str, Any], aliases: list[str], default: Any = "") -> Any:
    normalized_aliases = [_normalize_key(alias) for alias in aliases]
    for alias in normalized_aliases:
        for key, value in record.items():
            if key == alias:
                return value
    for alias in sorted(normalized_aliases, key=len, reverse=True):
        for key, value in record.items():
            if alias in key:
                return value
    return default


def _has_field_group(fields: set[str], group: str) -> bool:
    aliases = {_normalize_key(alias) for alias in FIELD_GROUP_ALIASES[group]}
    return any(field in aliases or any(alias in field for alias in aliases) for field in fields)


def _record_value(record: dict[str, Any], group: str, default: Any = "") -> Any:
    return _find_value(record, FIELD_GROUP_ALIASES[group], default)


def _sum_group(rows: list[dict[str, Any]], group: str) -> float:
    return sum(_as_float(_record_value(row, group, 0)) for row in rows)


def _sum_abs_group(rows: list[dict[str, Any]], group: str) -> float:
    return sum(abs(_as_float(_record_value(row, group, 0))) for row in rows)


def _non_empty_rate(rows: list[dict[str, Any]], field: str) -> float:
    if not rows:
        return 0.0
    non_empty = sum(1 for row in rows if row.get(field) not in {"", None})
    return non_empty / len(rows)


def _as_float(value: Any, default: float = 0.0) -> float:
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace("%", "").strip())
    except (TypeError, ValueError):
        return default


def _sku_matches(record: dict[str, Any], sku: str) -> bool:
    target = sku.strip().lower()
    values = [
        _find_value(record, ["sku", "msku", "seller_sku", "asin", "商品编码", "货号"]),
        _find_value(record, ["product_name", "name", "title", "商品名称", "产品名称"]),
    ]
    return any(target in str(value).strip().lower() for value in values if value)


@dataclass
class SkuSnapshot:
    sku: str
    product_name: str
    current_stock: float
    inbound_stock: float
    daily_sales_7d: float
    daily_sales_30d: float
    lead_time_days: float
    unit_cost: float
    selling_price: float
    gross_margin: float
    acos_7d: float
    roas_7d: float

    @property
    def total_available(self) -> float:
        return self.current_stock + self.inbound_stock

    @property
    def effective_daily_sales(self) -> float:
        return self.daily_sales_7d or self.daily_sales_30d or 1.0

    @property
    def days_cover(self) -> float:
        return self.current_stock / self.effective_daily_sales

    @property
    def gross_profit_per_unit(self) -> float:
        if self.gross_margin:
            return self.selling_price * self.gross_margin
        return self.selling_price - self.unit_cost

    def to_dict(self) -> dict[str, Any]:
        return {
            "sku": self.sku,
            "product_name": self.product_name,
            "current_stock": round(self.current_stock, 2),
            "inbound_stock": round(self.inbound_stock, 2),
            "daily_sales_7d": round(self.daily_sales_7d, 2),
            "daily_sales_30d": round(self.daily_sales_30d, 2),
            "lead_time_days": round(self.lead_time_days, 2),
            "days_cover": round(self.days_cover, 1),
            "unit_cost": round(self.unit_cost, 2),
            "selling_price": round(self.selling_price, 2),
            "gross_profit_per_unit": round(self.gross_profit_per_unit, 2),
            "acos_7d": round(self.acos_7d, 2),
            "roas_7d": round(self.roas_7d, 2),
        }


def _snapshot_from_fact_layer(sku: str) -> SkuSnapshot | None:
    if not _has_fact_table("fact_inventory_daily") and not _has_fact_table("fact_sales_daily"):
        return None

    inventory_rows = []
    sales_rows = []
    if _has_fact_table("fact_inventory_daily"):
        inventory_rows = _query_fact_rows(
            f"""
            WITH base AS (
                SELECT *
                FROM marts.fact_inventory_daily
                WHERE LOWER(COALESCE(sku, '')) = LOWER({_sql_string(sku)})
            ),
            latest_date AS (
                SELECT MAX(date_value) AS max_date
                FROM base
                WHERE date_value IS NOT NULL
            ),
            latest AS (
                SELECT *
                FROM base
                WHERE date_value IS NULL
                   OR date_value = (SELECT max_date FROM latest_date)
                   OR (SELECT max_date FROM latest_date) IS NULL
            )
            SELECT
                ANY_VALUE(product_name) AS product_name,
                SUM(COALESCE(ending_inventory, 0)) AS current_stock,
                SUM(COALESCE(in_transit, 0)) + SUM(COALESCE(inbound, 0)) AS inbound_stock
            FROM latest
            """,
            limit=5,
        )
    if _has_fact_table("fact_sales_daily"):
        sales_rows = _query_fact_rows(
            f"""
            WITH base AS (
                SELECT
                    COALESCE(date_value, current_date) AS metric_date,
                    COALESCE(sales_qty, 0) AS sales_qty
                FROM marts.fact_sales_daily
                WHERE LOWER(COALESCE(sku, '')) = LOWER({_sql_string(sku)})
            )
            SELECT
                AVG(sales_qty) FILTER (WHERE metric_date >= current_date - INTERVAL 7 DAY) AS daily_sales_7d,
                AVG(sales_qty) FILTER (WHERE metric_date >= current_date - INTERVAL 30 DAY) AS daily_sales_30d
            FROM base
            """,
            limit=5,
        )
    inventory = inventory_rows[0] if inventory_rows else {}
    sales = sales_rows[0] if sales_rows else {}
    if not inventory and not sales:
        return None
    product_name = str(inventory.get("product_name") or sku)
    return SkuSnapshot(
        sku=sku,
        product_name=product_name,
        current_stock=_as_float(inventory.get("current_stock", 0)),
        inbound_stock=_as_float(inventory.get("inbound_stock", 0)),
        daily_sales_7d=_as_float(sales.get("daily_sales_7d", 0)),
        daily_sales_30d=_as_float(sales.get("daily_sales_30d", 0)),
        lead_time_days=30.0,
        unit_cost=0.0,
        selling_price=0.0,
        gross_margin=0.0,
        acos_7d=0.0,
        roas_7d=0.0,
    )


def _snapshot_for_sku(sku: str) -> SkuSnapshot:
    fact_snapshot = _snapshot_from_fact_layer(sku)
    if fact_snapshot is not None:
        return fact_snapshot

    data = _load_business_data()
    matching_records: list[dict[str, Any]] = []
    for rows in data.values():
        matching_records.extend([row for row in rows if _sku_matches(row, sku)])

    inventory = next((row for row in data.get("inventory", []) if _sku_matches(row, sku)), {})
    sales = next((row for row in data.get("sales", []) if _sku_matches(row, sku)), {})
    ads = next((row for row in data.get("ads", []) if _sku_matches(row, sku)), {})
    profit = next((row for row in data.get("profit", []) if _sku_matches(row, sku)), {})
    product = next((row for row in data.get("products", []) if _sku_matches(row, sku)), {})
    fallback = matching_records[0] if matching_records else {}

    product_name = str(
        _find_value(product, ["product_name", "name", "title", "商品名称"], "")
        or _find_value(inventory, ["product_name", "name", "title", "商品名称"], "")
        or _find_value(fallback, ["product_name", "name", "title", "商品名称"], "")
        or sku
    )

    return SkuSnapshot(
        sku=str(_find_value(fallback, ["sku", "msku", "seller_sku", "asin", "商品编码"], sku)),
        product_name=product_name,
        current_stock=_as_float(_find_value(inventory, ["current_stock", "stock", "available", "可售库存", "库存"], 0)),
        inbound_stock=_as_float(_find_value(inventory, ["inbound_stock", "inbound", "on_the_way", "在途", "在途库存"], 0)),
        daily_sales_7d=_as_float(_find_value(sales, ["daily_sales_7d", "avg_daily_sales_7d", "7d_daily_sales", "近7日日均销量"], 0)),
        daily_sales_30d=_as_float(_find_value(sales, ["daily_sales_30d", "avg_daily_sales_30d", "30d_daily_sales", "近30日日均销量"], 0)),
        lead_time_days=_as_float(_find_value(inventory, ["lead_time_days", "purchase_lead_time", "补货周期", "采购周期"], 30)),
        unit_cost=_as_float(_find_value(profit, ["unit_cost", "cost", "采购成本", "单件成本"], 0)),
        selling_price=_as_float(_find_value(profit, ["selling_price", "price", "售价", "销售价"], 0)),
        gross_margin=_as_float(_find_value(profit, ["gross_margin", "margin", "毛利率"], 0)),
        acos_7d=_as_float(_find_value(ads, ["acos_7d", "acos", "7d_acos"], 0)),
        roas_7d=_as_float(_find_value(ads, ["roas_7d", "roas", "7d_roas"], 0)),
    )


def list_registered_datasets() -> str:
    """列出已经进入 DuckDB 事实层注册表的数据集。"""
    return list_registered_datasets_from_registry()


def register_all_fact_datasets() -> str:
    """把历史大表 manifest 和标准结构化文件统一注册进 DuckDB 事实层。"""
    return register_all_fact_datasets_from_registry()


def audit_fact_source_readiness() -> str:
    """检查本地结构化文件能否直接长出 inventory/sales/finance/ads fact marts。"""
    return audit_fact_source_readiness_from_registry()


def summarize_brand_coverage(limit: int = 100) -> str:
    """汇总当前事实层里能识别到的品牌范围，帮助 agent 避免默认绑定单一品牌。"""
    if _has_fact_table("dim_product_master"):
        columns = _query_fact_rows(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_schema = 'marts' AND table_name = 'dim_product_master'",
            limit=50,
        )
        column_names = {str(item.get("column_name", "")) for item in columns}
        if "brand" not in column_names:
            return json.dumps(
                {
                    "row_count": 0,
                    "rows": [],
                    "warning": "dim_product_master exists but does not expose a brand column yet; re-register the fact layer.",
                },
                ensure_ascii=False,
                indent=2,
            )
        return query_fact_layer(
            "SELECT dataset_slug, brand, COUNT(*) AS sku_count "
            "FROM marts.dim_product_master "
            "WHERE brand IS NOT NULL AND TRIM(brand) <> '' "
            "GROUP BY 1, 2 "
            "ORDER BY sku_count DESC NULLS LAST, dataset_slug, brand",
            limit=limit,
        )
    data = _load_business_data()
    brand_like_fields = ["brand", "品牌", "品牌名称"]
    rows = []
    for dataset_name, dataset_rows in data.items():
        counter: dict[str, int] = {}
        for row in dataset_rows:
            for field in brand_like_fields:
                value = row.get(_normalize_key(field), "")
                text = str(value).strip()
                if text:
                    counter[text] = counter.get(text, 0) + 1
        for brand, sku_count in sorted(counter.items(), key=lambda item: (-item[1], item[0]))[:limit]:
            rows.append({"dataset_slug": dataset_name, "brand": brand, "sku_count": sku_count})
    return json.dumps({"row_count": len(rows), "rows": rows[:limit]}, ensure_ascii=False, indent=2)


def list_fact_tables() -> str:
    """列出 DuckDB 数据集视图和 mart 视图。"""
    return list_fact_tables_from_registry()


def query_fact_layer(sql: str, limit: int = 200) -> str:
    """执行只读 DuckDB 查询，适合全量结构化指标分析。

    只允许查询 datasets.* 和 marts.*。如需看表结构或可用对象，先调用 list_fact_tables；
    不要查询 information_schema、系统表、外部文件或裸 DuckDB 表名。
    """
    try:
        return query_fact_layer_from_registry(sql, limit=limit)
    except ValueError as exc:
        return json.dumps(
            {
                "status": "error",
                "available": False,
                "error_type": "fact_layer_policy_violation",
                "error": str(exc),
                "row_count": 0,
                "rows": [],
                "allowed_object_prefixes": ["datasets.*", "marts.*"],
                "recovery_hint": (
                    "Use list_fact_tables to inspect available fact-layer objects, then query only "
                    "datasets.<view_name> or marts.<view_name>. information_schema and bare table names "
                    "are intentionally blocked."
                ),
            },
            ensure_ascii=False,
            indent=2,
        )


def query_inventory_history(sku: str = "", days: int = 30, warehouse: str = "", limit: int = 200) -> str:
    """查询 DuckDB 事实层中的库存明细，优先用于大表全量分析。"""
    if not _has_fact_table("fact_inventory_daily"):
        return json.dumps(
            {
                "available": False,
                "reason": "fact_inventory_daily mart not available",
                "fallback": "use warehouse manifest/quality or sampled CSV analysis",
            },
            ensure_ascii=False,
            indent=2,
        )
    nonzero_first = "CASE WHEN COALESCE(ending_inventory, 0) <> 0 THEN 0 ELSE 1 END"
    filters = []
    if sku.strip():
        filters.append(f"LOWER(COALESCE(sku, '')) = LOWER({_sql_string(sku.strip())})")
    if warehouse.strip():
        filters.append(f"LOWER(COALESCE(warehouse, '')) = LOWER({_sql_string(warehouse.strip())})")
    if days > 0:
        filters.append(
            "(effective_date_value IS NULL OR effective_date_value >= "
            f"(SELECT COALESCE(MAX(effective_date_value), current_date) FROM marts.fact_inventory_daily) - INTERVAL {int(days)} DAY)"
        )
    where = f"WHERE {' AND '.join(filters)}" if filters else ""
    sql = (
        "SELECT dataset_slug, effective_date_value AS date_value, snapshot_date, date_text, sku, product_name, warehouse, "
        "opening_inventory, inbound, outbound, ending_inventory, in_transit "
        "FROM marts.fact_inventory_daily "
        f"{where} "
        f"ORDER BY {nonzero_first}, effective_date_value DESC NULLS LAST, ABS(COALESCE(ending_inventory, 0)) DESC, sku, warehouse"
    )
    return query_fact_layer(sql, limit=limit)


def query_inventory_snapshot(sku: str = "", warehouse: str = "", nonzero_only: bool = False, limit: int = 200) -> str:
    """查询当前库存快照，优先返回非零库存和最近快照。"""
    if not _has_fact_table("current_inventory_snapshot"):
        return json.dumps(
            {
                "available": False,
                "reason": "current_inventory_snapshot view not available",
                "fallback": "use query_inventory_history",
            },
            ensure_ascii=False,
            indent=2,
        )
    filters = []
    if sku.strip():
        filters.append(f"LOWER(COALESCE(sku, '')) = LOWER({_sql_string(sku.strip())})")
    if warehouse.strip():
        filters.append(f"LOWER(COALESCE(warehouse, '')) = LOWER({_sql_string(warehouse.strip())})")
    if nonzero_only:
        filters.append("COALESCE(ending_inventory, 0) <> 0")
    where = f"WHERE {' AND '.join(filters)}" if filters else ""
    sql = (
        "SELECT dataset_slug, effective_date_value AS date_value, snapshot_date, sku, product_name, warehouse, "
        "opening_inventory, inbound, outbound, ending_inventory, in_transit "
        "FROM marts.current_inventory_snapshot "
        f"{where} "
        "ORDER BY CASE WHEN COALESCE(ending_inventory, 0) <> 0 THEN 0 ELSE 1 END, ABS(COALESCE(ending_inventory, 0)) DESC, sku, warehouse"
    )
    return query_fact_layer(sql, limit=limit)


def _inventory_anomaly_types_from_question(text: str) -> list[str]:
    lowered = text.lower()
    anomaly_types = []
    rules = [
        ("negative_inventory", ["负库存", "库存为负", "negative"]),
        ("inbound_outbound_imbalance", ["出入库不平", "出入库不一致", "平衡异常", "不平", "不一致", "imbalance"]),
        ("no_sales_30d_high_inventory", ["无销量高库存", "无销量", "零销量", "滞销", "高库存", "积压"]),
        ("stockout_risk", ["断货", "缺货", "售罄", "stockout"]),
    ]
    for anomaly_type, tokens in rules:
        if any(token in lowered for token in tokens):
            anomaly_types.append(anomaly_type)
    if not anomaly_types and any(token in lowered for token in ["异常", "风险", "库存问题"]):
        anomaly_types = [rule[0] for rule in rules]
    return list(dict.fromkeys(anomaly_types))


def query_inventory_anomalies(anomaly_type: str = "", sku: str = "", warehouse: str = "", limit: int = 200) -> str:
    """查询库存异常 mart：负库存、出入库不平、30 天无销量高库存、断货风险。"""
    if not _has_fact_table("inventory_anomalies"):
        return json.dumps(
            {
                "available": False,
                "reason": "inventory_anomalies mart not available",
                "fallback": "register the fact layer after inventory and sales data are available",
            },
            ensure_ascii=False,
            indent=2,
        )
    filters = []
    if anomaly_type.strip():
        filters.append(f"anomaly_type = {_sql_string(anomaly_type.strip())}")
    if sku.strip():
        filters.append(f"LOWER(COALESCE(sku, '')) = LOWER({_sql_string(sku.strip())})")
    if warehouse.strip():
        filters.append(f"LOWER(COALESCE(warehouse, '')) = LOWER({_sql_string(warehouse.strip())})")
    where = f"WHERE {' AND '.join(filters)}" if filters else ""
    sql = (
        "SELECT anomaly_type, severity, dataset_slug, effective_date_value AS date_value, sku, sku_hash_bucket, "
        "product_name, warehouse, ending_inventory, inbound, outbound, opening_inventory, in_transit, "
        "sales_qty_30d, selling_days_30d, recommendation "
        "FROM marts.inventory_anomalies "
        f"{where} "
        "ORDER BY CASE severity WHEN 'high' THEN 0 WHEN 'medium' THEN 1 ELSE 2 END, anomaly_type, ABS(COALESCE(ending_inventory, 0)) DESC, sku"
    )
    return query_fact_layer(sql, limit=limit)


def query_sales_history(sku: str = "", days: int = 30, warehouse: str = "", limit: int = 200) -> str:
    """查询 DuckDB 事实层中的销量明细，优先用于大表全量分析。"""
    if not _has_fact_table("fact_sales_daily"):
        return json.dumps(
            {
                "available": False,
                "reason": "fact_sales_daily mart not available",
                "fallback": "use sampled CSV analysis",
            },
            ensure_ascii=False,
            indent=2,
        )
    filters = []
    if sku.strip():
        filters.append(f"LOWER(COALESCE(sku, '')) = LOWER({_sql_string(sku.strip())})")
    if warehouse.strip():
        filters.append(f"LOWER(COALESCE(warehouse, '')) = LOWER({_sql_string(warehouse.strip())})")
    if days > 0:
        filters.append(
            "(effective_date_value IS NULL OR effective_date_value >= "
            f"(SELECT COALESCE(MAX(effective_date_value), current_date) FROM marts.fact_sales_daily) - INTERVAL {int(days)} DAY)"
        )
    where = f"WHERE {' AND '.join(filters)}" if filters else ""
    sql = (
        "SELECT dataset_slug, effective_date_value AS date_value, snapshot_date, date_text, sku, product_name, warehouse, channel, sales_qty, revenue "
        "FROM marts.fact_sales_daily "
        f"{where} "
        "ORDER BY effective_date_value DESC NULLS LAST, sku, warehouse, channel"
    )
    return query_fact_layer(sql, limit=limit)


def query_finance_history(sku: str = "", days: int = 30, warehouse: str = "", channel: str = "", limit: int = 200) -> str:
    """查询 DuckDB 事实层中的财务明细，优先用于收入/成本/毛利/现金分析。"""
    if not _has_fact_table("fact_finance_daily"):
        return json.dumps(
            {
                "available": False,
                "reason": "fact_finance_daily mart not available",
                "fallback": "use local sampled profit/finance tables when present",
            },
            ensure_ascii=False,
            indent=2,
        )
    filters = []
    if sku.strip():
        filters.append(f"LOWER(COALESCE(sku, '')) = LOWER({_sql_string(sku.strip())})")
    if warehouse.strip():
        filters.append(f"LOWER(COALESCE(warehouse, '')) = LOWER({_sql_string(warehouse.strip())})")
    if channel.strip():
        filters.append(f"LOWER(COALESCE(channel, '')) = LOWER({_sql_string(channel.strip())})")
    if days > 0:
        filters.append(
            "(effective_date_value IS NULL OR effective_date_value >= "
            f"(SELECT COALESCE(MAX(effective_date_value), current_date) FROM marts.fact_finance_daily) - INTERVAL {int(days)} DAY)"
        )
    where = f"WHERE {' AND '.join(filters)}" if filters else ""
    sql = (
        "SELECT dataset_slug, effective_date_value AS date_value, snapshot_date, date_text, sku, product_name, warehouse, channel, revenue, cost, gross_profit, cash "
        "FROM marts.fact_finance_daily "
        f"{where} "
        "ORDER BY effective_date_value DESC NULLS LAST, sku, warehouse, channel"
    )
    return query_fact_layer(sql, limit=limit)


def query_ads_history(sku: str = "", days: int = 30, warehouse: str = "", channel: str = "", limit: int = 200) -> str:
    """查询 DuckDB 事实层中的广告明细，优先用于广告花费/ACOS/ROAS 分析。"""
    if not _has_fact_table("fact_ads_daily"):
        return json.dumps(
            {
                "available": False,
                "reason": "fact_ads_daily mart not available",
                "fallback": "use local sampled ads tables when present",
            },
            ensure_ascii=False,
            indent=2,
        )
    filters = []
    if sku.strip():
        filters.append(f"LOWER(COALESCE(sku, '')) = LOWER({_sql_string(sku.strip())})")
    if warehouse.strip():
        filters.append(f"LOWER(COALESCE(warehouse, '')) = LOWER({_sql_string(warehouse.strip())})")
    if channel.strip():
        filters.append(f"LOWER(COALESCE(channel, '')) = LOWER({_sql_string(channel.strip())})")
    if days > 0:
        filters.append(
            "(effective_date_value IS NULL OR effective_date_value >= "
            f"(SELECT COALESCE(MAX(effective_date_value), current_date) FROM marts.fact_ads_daily) - INTERVAL {int(days)} DAY)"
        )
    where = f"WHERE {' AND '.join(filters)}" if filters else ""
    sql = (
        "SELECT dataset_slug, effective_date_value AS date_value, snapshot_date, date_text, sku, product_name, warehouse, channel, ad_spend, acos, roas "
        "FROM marts.fact_ads_daily "
        f"{where} "
        "ORDER BY effective_date_value DESC NULLS LAST, sku, warehouse, channel"
    )
    return query_fact_layer(sql, limit=limit)


def plan_fact_query(question: str, limit: int = 200) -> str:
    """把自然语言问题映射成受控 DuckDB 查询计划，不直接接受自由 SQL。"""
    text = question.strip()
    lowered = text.lower()
    table = ""
    query_kind = "history"
    group_by: list[str] = []
    anomaly_types = _inventory_anomaly_types_from_question(text)
    if anomaly_types:
        table = "inventory_anomalies"
        query_kind = "anomaly"
    if any(token in lowered for token in ["广告", "acos", "roas", "投放", "ad "]):
        table = "fact_ads_daily"
    elif any(token in lowered for token in ["收入", "成本", "毛利", "现金", "利润", "财务", "revenue", "profit", "cash"]):
        table = "fact_finance_daily"
    elif any(token in lowered for token in ["销量", "销售", "出货", "gmv", "order", "订单"]):
        table = "fact_sales_daily"
    elif any(token in lowered for token in ["库存", "在途", "补货", "仓库", "断货", "积压"]):
        table = "fact_inventory_daily"
    if anomaly_types:
        table = "inventory_anomalies"
        query_kind = "anomaly"

    if table == "fact_inventory_daily" and any(token in lowered for token in ["快照", "当前", "现在", "可售库存", "库存现状"]):
        query_kind = "snapshot"
    elif any(token in lowered for token in ["汇总", "总和", "总计", "合计", "sum", "总览"]):
        query_kind = "aggregate"
    elif any(token in lowered for token in ["排名", "top", "最高", "最多", "最大", "前"]) and table in {"fact_inventory_daily", "fact_sales_daily", "fact_ads_daily"}:
        query_kind = "top"
    if table == "fact_sales_daily" and any(token in lowered for token in ["对比", "环比", "较", "本周", "上周", "本期", "上期"]):
        query_kind = "time_compare"
        table = "agg_sku_daily_sales"
    elif table == "fact_sales_daily" and any(token in lowered for token in ["按渠道", "分渠道", "渠道汇总", "渠道销量"]):
        query_kind = "group"
        group_by = ["channel"]
        table = "agg_channel_sales"
    elif table == "fact_sales_daily" and any(token in lowered for token in ["按sku", "按 sku", "sku汇总", "sku销量"]):
        query_kind = "group"
        group_by = ["sku"]
        table = "agg_sku_daily_sales"
    elif table == "fact_sales_daily" and query_kind == "top":
        table = "agg_sku_daily_sales"
    elif table == "fact_inventory_daily" and any(token in lowered for token in ["按仓库", "仓库汇总", "仓库库存"]):
        query_kind = "group"
        group_by = ["warehouse"]
        table = "agg_warehouse_inventory"

    days = 30
    day_match = re.search(r"(最近|近)(\d{1,3})\s*天", text)
    if day_match:
        days = int(day_match.group(2))
    elif "今天" in text:
        days = 1
    elif "最近7天" in text or "近7天" in text:
        days = 7

    sku_match = re.search(r"\b[A-Z0-9][A-Z0-9_-]{3,}\b", text)
    sku = sku_match.group(0) if sku_match else ""
    warehouse = ""
    warehouse_match = re.search(r"([\u4e00-\u9fffA-Za-z0-9_-]{2,}(?:仓|仓库))", text)
    if warehouse_match:
        warehouse = warehouse_match.group(1)
    channel = ""
    channel_terms = ["天猫国际", "天猫", "淘宝", "抖音", "拼多多", "京东", "快手", "小红书", "唯品会", "得物"]
    for term in channel_terms:
        if term in lowered:
            channel = term
            break
    top_n = 20
    top_match = re.search(r"(top|前)\s*(\d{1,3})", lowered)
    if top_match:
        top_n = int(top_match.group(2))

    plan: dict[str, Any] = {
        "question": question,
        "table": table,
        "query_kind": query_kind,
        "days": days,
        "sku": sku,
        "warehouse": warehouse,
        "channel": channel,
        "group_by": group_by,
        "anomaly_types": anomaly_types,
        "limit": max(1, min(limit, 500)),
        "top_n": max(1, min(top_n, 100)),
        "available": False,
        "sql": "",
        "reason": "",
    }
    if not table:
        plan["reason"] = "Could not safely map the question to a known fact mart."
        return json.dumps(plan, ensure_ascii=False, indent=2)
    required_table = "current_inventory_snapshot" if table == "fact_inventory_daily" and query_kind == "snapshot" else table
    if not _has_fact_table(required_table):
        plan["reason"] = f"{table} mart is not available."
        return json.dumps(plan, ensure_ascii=False, indent=2)

    filters = []
    if sku:
        filters.append(f"LOWER(COALESCE(sku, '')) = LOWER({_sql_string(sku)})")
    if warehouse and table not in {"agg_sku_daily_sales", "agg_channel_sales"}:
        filters.append(f"LOWER(COALESCE(warehouse, '')) LIKE LOWER({_sql_string('%' + warehouse + '%')})")
    if channel and table != "agg_sku_daily_sales":
        filters.append(f"LOWER(COALESCE(channel, '')) LIKE LOWER({_sql_string('%' + channel + '%')})")
    if days > 0 and query_kind not in {"snapshot", "time_compare", "anomaly"} and table != "agg_channel_sales":
        if table == "agg_sku_daily_sales":
            filters.append(
                f"(date_value IS NULL OR date_value >= (SELECT MAX(date_value) FROM marts.agg_sku_daily_sales WHERE date_value IS NOT NULL) - INTERVAL {int(days)} DAY)"
            )
        elif table == "agg_warehouse_inventory":
            filters.append(
                f"(snapshot_date IS NULL OR snapshot_date >= (SELECT MAX(snapshot_date) FROM marts.agg_warehouse_inventory WHERE snapshot_date IS NOT NULL) - INTERVAL {int(days)} DAY)"
            )
        else:
            filters.append(
                f"(effective_date_value IS NULL OR effective_date_value >= (SELECT MAX(effective_date_value) FROM marts.{table} WHERE effective_date_value IS NOT NULL) - INTERVAL {int(days)} DAY)"
            )
    if table == "inventory_anomalies" and anomaly_types:
        filters.append("anomaly_type IN (" + ", ".join(_sql_string(item) for item in anomaly_types) + ")")
    if table in {"agg_sku_daily_sales", "agg_channel_sales"}:
        filters = [item.replace("effective_date_value", "date_value") for item in filters]
    if table == "agg_warehouse_inventory":
        filters = [item.replace("effective_date_value", "snapshot_date") for item in filters]
    where = f"WHERE {' AND '.join(filters)}" if filters else ""

    if table == "agg_sku_daily_sales" and query_kind == "time_compare":
        comparison_filters = []
        if sku:
            comparison_filters.append(f"LOWER(COALESCE(sku, '')) = LOWER({_sql_string(sku)})")
        comparison_filters.append("date_value >= (SELECT week_start FROM anchor) - INTERVAL 7 DAY")
        comparison_where = f"WHERE {' AND '.join(comparison_filters)}"
        sql = (
            "WITH anchor AS ("
            "SELECT MAX(date_value) AS max_date, DATE_TRUNC('week', MAX(date_value)) AS week_start "
            "FROM marts.agg_sku_daily_sales WHERE date_value IS NOT NULL"
            ") "
            "SELECT sku, "
            "ANY_VALUE(product_name) FILTER (WHERE product_name IS NOT NULL) AS product_name, "
            "SUM(CASE WHEN date_value >= (SELECT week_start FROM anchor) THEN COALESCE(sales_qty, 0) ELSE 0 END) AS current_period_sales_qty, "
            "SUM(CASE WHEN date_value >= (SELECT week_start FROM anchor) - INTERVAL 7 DAY AND date_value < (SELECT week_start FROM anchor) THEN COALESCE(sales_qty, 0) ELSE 0 END) AS previous_period_sales_qty, "
            "SUM(CASE WHEN date_value >= (SELECT week_start FROM anchor) THEN COALESCE(sales_qty, 0) ELSE 0 END) - "
            "SUM(CASE WHEN date_value >= (SELECT week_start FROM anchor) - INTERVAL 7 DAY AND date_value < (SELECT week_start FROM anchor) THEN COALESCE(sales_qty, 0) ELSE 0 END) AS sales_qty_delta "
            "FROM marts.agg_sku_daily_sales "
            f"{comparison_where} "
            "GROUP BY sku "
            "ORDER BY ABS(sales_qty_delta) DESC NULLS LAST, sku"
        )
    elif table == "agg_sku_daily_sales" and query_kind == "top":
        sql = (
            "SELECT sku, ANY_VALUE(product_name) FILTER (WHERE product_name IS NOT NULL) AS product_name, "
            "SUM(sales_qty) AS sales_qty, SUM(revenue) AS revenue "
            "FROM marts.agg_sku_daily_sales "
            f"{where} "
            "GROUP BY sku "
            f"ORDER BY sales_qty DESC NULLS LAST, revenue DESC NULLS LAST LIMIT {plan['top_n']}"
        )
    elif table == "agg_sku_daily_sales" and query_kind == "group":
        sql = (
            "SELECT sku, ANY_VALUE(product_name) FILTER (WHERE product_name IS NOT NULL) AS product_name, "
            "SUM(sales_qty) AS sales_qty, SUM(revenue) AS revenue "
            "FROM marts.agg_sku_daily_sales "
            f"{where} "
            "GROUP BY sku "
            "ORDER BY sales_qty DESC NULLS LAST, revenue DESC NULLS LAST"
        )
    elif table == "agg_channel_sales" and query_kind == "group":
        sql = (
            "SELECT channel, SUM(sales_qty) AS sales_qty, SUM(revenue) AS revenue "
            "FROM marts.agg_channel_sales "
            f"{where} "
            "GROUP BY channel "
            "ORDER BY sales_qty DESC NULLS LAST, revenue DESC NULLS LAST"
        )
    elif table == "agg_warehouse_inventory" and query_kind == "group":
        sql = (
            "SELECT warehouse, SUM(sku_count) AS sku_count, SUM(opening_inventory) AS opening_inventory, "
            "SUM(inbound) AS inbound, SUM(outbound) AS outbound, SUM(ending_inventory) AS ending_inventory, SUM(in_transit) AS in_transit "
            "FROM marts.agg_warehouse_inventory "
            f"{where} "
            "GROUP BY warehouse "
            "ORDER BY ending_inventory DESC NULLS LAST, warehouse"
        )
    elif table == "inventory_anomalies":
        sql = (
            "SELECT anomaly_type, severity, dataset_slug, effective_date_value AS date_value, effective_month, sku, sku_hash_bucket, "
            "product_name, warehouse, ending_inventory, inbound, outbound, opening_inventory, in_transit, "
            "sales_qty_30d, selling_days_30d, recommendation "
            "FROM marts.inventory_anomalies "
            f"{where} "
            "ORDER BY CASE severity WHEN 'high' THEN 0 WHEN 'medium' THEN 1 ELSE 2 END, anomaly_type, ABS(COALESCE(ending_inventory, 0)) DESC, sku"
        )
    elif table == "fact_inventory_daily" and query_kind == "snapshot":
        nonzero_only = any(token in lowered for token in ["非零", "有货", "可售", "在库"])
        where_clauses = [clause for clause in filters]
        if nonzero_only:
            where_clauses.append("COALESCE(ending_inventory, 0) <> 0")
        snapshot_where = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
        sql = (
            "SELECT dataset_slug, effective_date_value AS date_value, snapshot_date, sku, product_name, warehouse, "
            "opening_inventory, inbound, outbound, ending_inventory, in_transit "
            "FROM marts.current_inventory_snapshot "
            f"{snapshot_where} "
            "ORDER BY CASE WHEN COALESCE(ending_inventory, 0) <> 0 THEN 0 ELSE 1 END, ABS(COALESCE(ending_inventory, 0)) DESC, sku, warehouse"
        )
    elif table == "fact_inventory_daily" and query_kind == "top":
        sql = (
            "SELECT dataset_slug, effective_date_value AS date_value, sku, product_name, warehouse, ending_inventory, inbound, outbound "
            "FROM marts.current_inventory_snapshot "
            f"{where} "
            f"ORDER BY ABS(COALESCE(ending_inventory, 0)) DESC, sku LIMIT {plan['top_n']}"
        )
    elif table == "fact_inventory_daily":
        sql = (
            "SELECT dataset_slug, effective_date_value AS date_value, snapshot_date, sku, product_name, warehouse, opening_inventory, inbound, outbound, ending_inventory, in_transit "
            "FROM marts.fact_inventory_daily "
            f"{where} "
            "ORDER BY CASE WHEN COALESCE(ending_inventory, 0) <> 0 THEN 0 ELSE 1 END, effective_date_value DESC NULLS LAST, ABS(COALESCE(ending_inventory, 0)) DESC, sku, warehouse"
        )
    elif table == "fact_sales_daily" and query_kind == "aggregate":
        sql = (
            "SELECT warehouse, channel, sku, SUM(sales_qty) AS sales_qty, SUM(revenue) AS revenue "
            "FROM marts.fact_sales_daily "
            f"{where} "
            "GROUP BY 1, 2, 3 "
            "ORDER BY sales_qty DESC NULLS LAST, revenue DESC NULLS LAST"
        )
    elif table == "fact_sales_daily" and query_kind == "top":
        sql = (
            "SELECT warehouse, channel, sku, product_name, SUM(sales_qty) AS sales_qty, SUM(revenue) AS revenue "
            "FROM marts.fact_sales_daily "
            f"{where} "
            "GROUP BY 1, 2, 3, 4 "
            f"ORDER BY sales_qty DESC NULLS LAST, revenue DESC NULLS LAST LIMIT {plan['top_n']}"
        )
    elif table == "fact_sales_daily":
        sql = (
            "SELECT dataset_slug, effective_date_value AS date_value, snapshot_date, sku, product_name, warehouse, channel, sales_qty, revenue "
            "FROM marts.fact_sales_daily "
            f"{where} "
            "ORDER BY effective_date_value DESC NULLS LAST, sku, warehouse, channel"
        )
    elif table == "fact_finance_daily" and query_kind == "aggregate":
        sql = (
            "SELECT effective_date_value AS date_value, sku, warehouse, channel, SUM(revenue) AS revenue, SUM(cost) AS cost, "
            "SUM(gross_profit) AS gross_profit, MAX(cash) AS cash "
            "FROM marts.fact_finance_daily "
            f"{where} "
            "GROUP BY 1, 2, 3, 4 "
            "ORDER BY date_value DESC NULLS LAST, gross_profit DESC NULLS LAST"
        )
    elif table == "fact_finance_daily":
        sql = (
            "SELECT dataset_slug, effective_date_value AS date_value, snapshot_date, sku, product_name, warehouse, channel, revenue, cost, gross_profit, cash "
            "FROM marts.fact_finance_daily "
            f"{where} "
            "ORDER BY effective_date_value DESC NULLS LAST, sku, warehouse, channel"
        )
    elif table == "fact_ads_daily" and query_kind in {"aggregate", "top"}:
        sql = (
            "SELECT effective_date_value AS date_value, sku, warehouse, channel, SUM(ad_spend) AS ad_spend, AVG(acos) AS acos, AVG(roas) AS roas "
            "FROM marts.fact_ads_daily "
            f"{where} "
            "GROUP BY 1, 2, 3, 4 "
            f"ORDER BY ad_spend DESC NULLS LAST, date_value DESC NULLS LAST LIMIT {plan['top_n'] if query_kind == 'top' else plan['limit']}"
        )
    else:
        sql = (
            "SELECT dataset_slug, effective_date_value AS date_value, snapshot_date, sku, product_name, warehouse, channel, ad_spend, acos, roas "
            "FROM marts.fact_ads_daily "
            f"{where} "
            "ORDER BY effective_date_value DESC NULLS LAST, sku, warehouse, channel"
        )

    plan["available"] = True
    plan["sql"] = sql
    return json.dumps(plan, ensure_ascii=False, indent=2)


def query_fact_layer_from_question(question: str, limit: int = 200) -> str:
    """受控自然语言查询入口：先规划，再执行 DuckDB 查询。"""
    plan = json.loads(plan_fact_query(question, limit=limit))
    if not plan.get("available"):
        return json.dumps(plan, ensure_ascii=False, indent=2)
    result = json.loads(query_fact_layer(plan["sql"], limit=plan["limit"]))
    return json.dumps({"plan": plan, "result": result}, ensure_ascii=False, indent=2)


def list_business_files() -> str:
    """列出 D:\\A2A\\data 目录里可供 Agent 读取的业务数据文件。"""
    _ensure_data_dir()
    fact_layer = _fact_layer_summary()
    files = [
        {
            "path": str(path.relative_to(DATA_DIR)).replace("\\", "/"),
            "name": path.name,
            "dataset": _infer_dataset_name(path),
            "size_bytes": path.stat().st_size,
            "load_policy": "warehouse_sample" if _is_warehouse_path(path) else "standard_capped",
            "max_rows_loaded": _row_limit_for(path),
        }
        for path in sorted(DATA_DIR.rglob("*"))
        if _should_include_business_file(path)
    ]
    return json.dumps(
        {
            "data_dir": str(DATA_DIR),
            "files": files,
            "fact_layer": fact_layer,
            "load_policy": {
                "standard_max_rows_per_file": MAX_ROWS_PER_DATA_FILE,
                "warehouse_max_rows_per_file": MAX_WAREHOUSE_ROWS_PER_FILE,
                "skipped_dirs": sorted(_SKIPPED_DATA_DIRS),
            },
        },
        ensure_ascii=False,
        indent=2,
    )


def summarize_business_data() -> str:
    """汇总本地库存、销量、广告、利润等业务数据表的规模和字段。"""
    data = _load_business_data()
    fact_layer = _fact_layer_summary()
    summary = {}
    for dataset, rows in data.items():
        fields = sorted({field for row in rows for field in row.keys()})
        summary[dataset] = {
            "rows": len(rows),
            "fields": fields[:30],
            "sample": rows[:2],
        }
    return json.dumps(
        {
            "summary": summary,
            "fact_layer": fact_layer,
            "load_policy": {
                "standard_max_rows_per_file": MAX_ROWS_PER_DATA_FILE,
                "warehouse_max_rows_per_file": MAX_WAREHOUSE_ROWS_PER_FILE,
                "note": "registered large datasets should be queried through DuckDB marts; raw warehouse CSV chunks remain a fallback/sample path only.",
            },
        },
        ensure_ascii=False,
        indent=2,
    )


def assess_data_quality(decision_goal: str = "company_decision") -> str:
    """检查公司级辅助决策前的数据质量、字段完整性和缺口。"""
    data = _load_business_data()
    fact_layer = _fact_layer_summary()
    required_by_goal = {
        "company_decision": ["sku", "product_name", "date", "inventory", "sales_qty", "revenue", "cost", "gross_profit", "cash", "ad_spend", "supplier"],
        "inventory_decision": ["sku", "product_name", "inventory", "inbound", "sales_qty", "lead_time", "cost"],
        "financial_decision": ["date", "revenue", "cost", "gross_profit", "cash", "ad_spend", "inventory"],
    }
    required_groups = required_by_goal.get(decision_goal, required_by_goal["company_decision"])
    dataset_reports = {}
    all_fields: set[str] = set()
    warnings = []
    if WAREHOUSE_DIR.exists() and any(WAREHOUSE_DIR.rglob("*.csv")):
        warnings.append(
            "检测到 data/warehouse 大表分块；常规业务分析只读取每个分块的可控样本，"
            "大表全量质量请以 manifest.json 和 quality_report.json 为准。"
        )
    if fact_layer.get("available") and fact_layer.get("datasets"):
        warnings.append(
            "检测到 DuckDB 事实层已注册的大表数据集；涉及全量库存/销量问题时，应优先查询 marts 视图而不是 warehouse_sample。"
        )

    for dataset, rows in data.items():
        fields = sorted({field for row in rows for field in row.keys()})
        field_set = set(fields)
        all_fields.update(field_set)
        empty_rates = {
            field: round(1 - _non_empty_rate(rows, field), 4)
            for field in fields[:80]
        }
        high_empty_fields = [
            field
            for field, empty_rate in empty_rates.items()
            if empty_rate >= 0.5
        ][:20]
        duplicate_key_count = 0
        seen_keys = set()
        for row in rows:
            key = "|".join(str(_record_value(row, group, "")) for group in ["sku", "date", "warehouse"])
            if key.strip("|") and key in seen_keys:
                duplicate_key_count += 1
            seen_keys.add(key)

        dataset_reports[dataset] = {
            "rows": len(rows),
            "fields": fields[:80],
            "matched_field_groups": [group for group in FIELD_GROUP_ALIASES if _has_field_group(field_set, group)],
            "high_empty_fields": high_empty_fields,
            "duplicate_sku_date_rows": duplicate_key_count,
            "sample": rows[:1],
        }
        if len(rows) == 0:
            warnings.append(f"{dataset} 数据集为空。")
        if high_empty_fields:
            warnings.append(f"{dataset} 存在高空值字段：{', '.join(high_empty_fields[:8])}。")
        if duplicate_key_count:
            warnings.append(f"{dataset} 存在 {duplicate_key_count} 行可能重复的 SKU+日期+仓库记录。")

    present_groups = [group for group in required_groups if _has_field_group(all_fields, group)]
    missing_groups = [group for group in required_groups if group not in present_groups]
    if missing_groups:
        warnings.append(f"公司级决策缺少关键字段组：{', '.join(missing_groups)}。")

    quality_level = "high"
    if missing_groups or warnings:
        quality_level = "medium"
    if len(missing_groups) >= 4 or not data:
        quality_level = "low"

    return json.dumps(
        {
            "decision_goal": decision_goal,
            "quality_level": quality_level,
            "datasets": dataset_reports,
            "fact_layer": fact_layer,
            "present_field_groups": present_groups,
            "missing_field_groups": missing_groups,
            "warnings": warnings,
            "recommendation": "可以进入辅助决策，但必须标注数据缺口。" if quality_level == "medium" else "数据质量不足，建议先补齐或清洗后再做公司级决策。" if quality_level == "low" else "数据质量较好，可以进入公司级辅助决策。",
        },
        ensure_ascii=False,
        indent=2,
    )


def analyze_company_financial_position() -> str:
    """基于本地业务数据做公司级财务和现金流初步分析。"""
    data = _load_business_data()
    all_rows = [row for rows in data.values() for row in rows]
    profit_rows = data.get("profit", [])
    sales_rows = data.get("sales", [])
    ads_rows = data.get("ads", [])
    inventory_rows = data.get("inventory", [])
    structured_revenue = 0.0
    structured_cost = 0.0
    structured_gross_profit = 0.0
    structured_cash = 0.0
    structured_ad_spend = 0.0
    structured_sales_units = 0.0
    structured_inventory_units = 0.0
    structured_inbound_units = 0.0
    if _has_fact_table("fact_sales_daily"):
        sales_fact = _query_fact_rows(
            """
            SELECT
                SUM(COALESCE(revenue, 0)) AS revenue,
                SUM(ABS(COALESCE(sales_qty, 0))) AS sales_units
            FROM marts.fact_sales_daily
            """,
            limit=1,
        )
        if sales_fact:
            structured_revenue = _as_float(sales_fact[0].get("revenue", 0))
            structured_sales_units = _as_float(sales_fact[0].get("sales_units", 0))
    if _has_fact_table("fact_finance_daily"):
        finance_fact = _query_fact_rows(
            """
            SELECT
                SUM(COALESCE(revenue, 0)) AS revenue,
                SUM(COALESCE(cost, 0)) AS cost,
                SUM(COALESCE(gross_profit, 0)) AS gross_profit,
                MAX(COALESCE(cash, 0)) AS cash
            FROM marts.fact_finance_daily
            """,
            limit=1,
        )
        if finance_fact:
            structured_revenue = _as_float(finance_fact[0].get("revenue", structured_revenue))
            structured_cost = _as_float(finance_fact[0].get("cost", 0))
            structured_gross_profit = _as_float(finance_fact[0].get("gross_profit", 0))
            structured_cash = _as_float(finance_fact[0].get("cash", 0))
    if _has_fact_table("fact_ads_daily"):
        ads_fact = _query_fact_rows(
            "SELECT SUM(COALESCE(ad_spend, 0)) AS ad_spend FROM marts.fact_ads_daily",
            limit=1,
        )
        if ads_fact:
            structured_ad_spend = _as_float(ads_fact[0].get("ad_spend", 0))
    if _has_fact_table("fact_inventory_daily"):
        inventory_fact = _query_fact_rows(
            """
            WITH latest_date AS (
                SELECT MAX(date_value) AS max_date
                FROM marts.fact_inventory_daily
                WHERE date_value IS NOT NULL
            ),
            latest AS (
                SELECT *
                FROM marts.fact_inventory_daily
                WHERE date_value IS NULL
                   OR date_value = (SELECT max_date FROM latest_date)
                   OR (SELECT max_date FROM latest_date) IS NULL
            )
            SELECT
                SUM(COALESCE(ending_inventory, 0)) AS inventory_units,
                SUM(COALESCE(inbound, 0)) + SUM(COALESCE(in_transit, 0)) AS inbound_units
            FROM latest
            """,
            limit=1,
        )
        if inventory_fact:
            structured_inventory_units = _as_float(inventory_fact[0].get("inventory_units", 0))
            structured_inbound_units = _as_float(inventory_fact[0].get("inbound_units", 0))

    revenue = structured_revenue or _sum_group(profit_rows, "revenue") or _sum_group(sales_rows, "revenue")
    cost = structured_cost or _sum_group(profit_rows, "cost")
    gross_profit = structured_gross_profit or _sum_group(profit_rows, "gross_profit") or (revenue - cost if revenue or cost else 0)
    cash = structured_cash or _sum_group(profit_rows, "cash") or _sum_group(all_rows, "cash")
    ad_spend = structured_ad_spend or _sum_group(ads_rows, "ad_spend") or _sum_group(profit_rows, "ad_spend")
    inventory_units = structured_inventory_units or _sum_group(inventory_rows, "inventory")
    inbound_units = structured_inbound_units or _sum_group(inventory_rows, "inbound")
    sales_units = structured_sales_units or _sum_abs_group(sales_rows, "sales_qty") or _sum_abs_group(inventory_rows, "sales_qty")
    inventory_cost_proxy = 0.0
    if inventory_rows:
        for row in inventory_rows:
            unit_cost = _as_float(_record_value(row, "cost", 0))
            inventory_qty = _as_float(_record_value(row, "inventory", 0))
            inventory_cost_proxy += unit_cost * inventory_qty

    gross_margin = gross_profit / revenue if revenue else 0
    ad_ratio = ad_spend / revenue if revenue else 0
    inventory_turnover_proxy = sales_units / inventory_units if inventory_units else 0
    risks = []
    if revenue <= 0:
        risks.append("缺少收入数据，无法可靠判断公司营收质量。")
    if cost <= 0:
        risks.append("缺少成本数据，毛利和现金压力判断不完整。")
    if cash <= 0:
        risks.append("缺少现金余额或可用资金数据，无法判断采购承压能力。")
    if inventory_units > 0 and inventory_turnover_proxy < 0.05:
        risks.append("库存出库/库存比例偏低，可能存在周转压力。")
    if ad_ratio > 0.25:
        risks.append("广告支出占收入比例偏高，需要检查投放效率。")
    if not risks:
        risks.append("未发现明显财务红线，但仍需结合现金流、应收应付和采购计划人工确认。")

    return json.dumps(
        {
            "metrics": {
                "revenue": round(revenue, 2),
                "cost": round(cost, 2),
                "gross_profit": round(gross_profit, 2),
                "gross_margin": round(gross_margin, 4),
                "cash": round(cash, 2),
                "ad_spend": round(ad_spend, 2),
                "ad_spend_to_revenue": round(ad_ratio, 4),
                "inventory_units": round(inventory_units, 2),
                "inbound_units": round(inbound_units, 2),
                "sales_units": round(sales_units, 2),
                "inventory_turnover_proxy": round(inventory_turnover_proxy, 4),
                "inventory_cost_proxy": round(inventory_cost_proxy, 2),
            },
            "risks": risks,
            "data_sources": list(data.keys()),
            "fact_layer_used": bool(structured_revenue or structured_cost or structured_gross_profit or structured_cash or structured_ad_spend or structured_sales_units or structured_inventory_units or structured_inbound_units),
            "notes": [
                "inventory_turnover_proxy 使用销量或出库量 / 期末库存估算。",
                "inventory_cost_proxy 只有在库存数据包含单位成本时才可靠。",
                "缺少应收、应付、税费、平台回款周期时，现金流结论只能作为辅助判断。",
            ],
        },
        ensure_ascii=False,
        indent=2,
    )


def analyze_company_strategy(focus: str = "next_month") -> str:
    """从公司经营视角汇总库存、财务、广告、供应链和数据缺口，给出策略建议。"""
    data = _load_business_data()
    quality = json.loads(assess_data_quality("company_decision"))
    finance = json.loads(analyze_company_financial_position())
    inventory_rows = data.get("inventory", [])
    structured_inventory_candidates = []
    if _has_fact_table("fact_inventory_daily"):
        structured_inventory_candidates = _query_fact_rows(
            """
            WITH latest_date AS (
                SELECT MAX(date_value) AS max_date
                FROM marts.fact_inventory_daily
                WHERE date_value IS NOT NULL
            ),
            latest_inventory AS (
                SELECT *
                FROM marts.fact_inventory_daily
                WHERE date_value IS NULL
                   OR date_value = (SELECT max_date FROM latest_date)
                   OR (SELECT max_date FROM latest_date) IS NULL
            ),
            sales_avg AS (
                SELECT
                    sku,
                    AVG(ABS(COALESCE(sales_qty, 0))) AS avg_daily_sales_30d
                FROM marts.fact_sales_daily
                WHERE date_value IS NULL
                   OR date_value >= current_date - INTERVAL 30 DAY
                GROUP BY sku
            )
            SELECT
                ANY_VALUE(i.sku) AS sku,
                ANY_VALUE(i.product_name) AS product_name,
                ANY_VALUE(i.warehouse) AS warehouse,
                SUM(COALESCE(i.ending_inventory, 0)) AS inventory,
                SUM(COALESCE(i.inbound, 0) + COALESCE(i.in_transit, 0)) AS inbound,
                COALESCE(MAX(s.avg_daily_sales_30d), 0) AS sales_qty
            FROM latest_inventory i
            LEFT JOIN sales_avg s ON i.sku = s.sku
            GROUP BY i.sku
            LIMIT 1000
            """,
            limit=1000,
        )

    stockout_candidates = []
    overstock_candidates = []
    source_rows = structured_inventory_candidates or inventory_rows[:1000]
    for row in source_rows:
        sku = row.get("sku", "") if structured_inventory_candidates else _record_value(row, "sku", "")
        name = row.get("product_name", "") if structured_inventory_candidates else _record_value(row, "product_name", "")
        inventory_qty = _as_float(row.get("inventory", 0) if structured_inventory_candidates else _record_value(row, "inventory", 0))
        inbound_qty = _as_float(row.get("inbound", 0) if structured_inventory_candidates else _record_value(row, "inbound", 0))
        sales_qty = abs(_as_float(row.get("sales_qty", 0) if structured_inventory_candidates else _record_value(row, "sales_qty", 0)))
        if sales_qty > 0 and inventory_qty / sales_qty < 7:
            stockout_candidates.append({"sku": sku, "name": name, "inventory": inventory_qty, "sales_qty": sales_qty, "inbound": inbound_qty})
        if inventory_qty > 0 and sales_qty == 0:
            overstock_candidates.append({"sku": sku, "name": name, "inventory": inventory_qty, "sales_qty": sales_qty, "inbound": inbound_qty})

    actions = []
    if quality["quality_level"] == "low":
        actions.append("先补齐核心经营数据，再做公司级动作：收入、成本、现金、库存、销量、广告、供应商。")
    if stockout_candidates:
        actions.append("优先复核断货候选 SKU，确认采购在途、真实销量周期和供应商交期。")
    if overstock_candidates:
        actions.append("梳理零动销或低动销库存，制定清库存、暂停采购或广告测试策略。")
    if finance["metrics"]["cash"] <= 0:
        actions.append("补充现金余额、应收应付和采购计划，否则不能判断公司承压能力。")
    if finance["metrics"]["ad_spend_to_revenue"] > 0.25:
        actions.append("广告预算先做结构优化，减少高花费低转化项。")
    if not actions:
        actions.append("现有数据未触发强红线，可进入产品线优先级、补货计划和广告预算分配讨论。")

    return json.dumps(
        {
            "focus": focus,
            "company_decision_readiness": quality["quality_level"],
            "key_data_gaps": quality["missing_field_groups"],
            "financial_snapshot": finance["metrics"],
            "fact_layer_used": bool(structured_inventory_candidates),
            "evidence_chain": _decision_evidence_chain(quality),
            "stockout_candidates": stockout_candidates[:20],
            "overstock_candidates": overstock_candidates[:20],
            "recommended_actions": actions,
            "decision_principles": [
                "先保证数据质量，再给出高金额经营建议。",
                "库存、现金流、广告效率和供应商交期必须交叉验证。",
                "公司级结论需要标注证据来源和人工确认项。",
            ],
        },
        ensure_ascii=False,
        indent=2,
    )


def query_sku_snapshot(sku: str) -> str:
    """查询单个 SKU 的库存、销量、广告和利润快照。"""
    snapshot = _snapshot_for_sku(sku)
    return json.dumps(snapshot.to_dict(), ensure_ascii=False, indent=2)


def analyze_restock_decision(sku: str, proposed_qty: int = 0) -> str:
    """基于本地数据评估一个 SKU 的补货建议，可传入计划补货数量。"""
    snapshot = _snapshot_for_sku(sku)
    demand_during_lead_time = snapshot.effective_daily_sales * snapshot.lead_time_days
    safety_stock = snapshot.effective_daily_sales * 14
    target_stock = demand_during_lead_time + safety_stock
    recommended_qty = max(0, math.ceil(target_stock - snapshot.total_available))
    qty = proposed_qty or recommended_qty
    stock_after_order = snapshot.total_available + qty
    projected_cover = stock_after_order / snapshot.effective_daily_sales
    cash_required = qty * snapshot.unit_cost
    gross_profit_potential = qty * snapshot.gross_profit_per_unit

    if snapshot.days_cover < snapshot.lead_time_days:
        stockout_risk = "high"
    elif snapshot.days_cover < snapshot.lead_time_days + 14:
        stockout_risk = "medium"
    else:
        stockout_risk = "low"

    overstock_risk = "high" if projected_cover > 120 else "medium" if projected_cover > 75 else "low"
    recommendation = "补货" if recommended_qty > 0 else "暂缓补货，继续观察"

    result = {
        "sku_snapshot": snapshot.to_dict(),
        "decision": {
            "recommendation": recommendation,
            "recommended_qty": recommended_qty,
            "proposed_qty": proposed_qty,
            "evaluated_qty": qty,
            "projected_days_cover_after_order": round(projected_cover, 1),
            "cash_required": round(cash_required, 2),
            "gross_profit_potential": round(gross_profit_potential, 2),
            "stockout_risk": stockout_risk,
            "overstock_risk": overstock_risk,
        },
        "assumptions": [
            "目标覆盖采购周期销量 + 14 天安全库存。",
            "若未传 proposed_qty，则使用系统测算的 recommended_qty。",
            "广告、利润和销量字段缺失时按 0 或保守默认值处理。",
        ],
    }
    return json.dumps(result, ensure_ascii=False, indent=2)


def simulate_decision_scenarios(sku: str, base_qty: int = 0) -> str:
    """生成保守、平衡、激进三个补货方案，并比较现金占用、覆盖天数和风险。"""
    snapshot = _snapshot_for_sku(sku)
    base_result = json.loads(analyze_restock_decision(sku, base_qty))
    recommended_qty = int(base_result["decision"]["recommended_qty"])
    if base_qty > 0:
        recommended_qty = base_qty

    scenarios = {
        "A_保守": max(0, math.floor(recommended_qty * 0.5)),
        "B_平衡": max(0, recommended_qty),
        "C_激进": max(0, math.ceil(recommended_qty * 1.5)),
    }
    rows = []
    for name, qty in scenarios.items():
        stock_after_order = snapshot.total_available + qty
        projected_cover = stock_after_order / snapshot.effective_daily_sales
        rows.append(
            {
                "scenario": name,
                "order_qty": qty,
                "cash_required": round(qty * snapshot.unit_cost, 2),
                "projected_days_cover": round(projected_cover, 1),
                "gross_profit_potential": round(qty * snapshot.gross_profit_per_unit, 2),
                "risk_note": "断货风险偏高" if projected_cover < snapshot.lead_time_days else "库存积压风险偏高" if projected_cover > 120 else "风险相对均衡",
            }
        )
    return json.dumps({"sku": snapshot.to_dict(), "scenarios": rows}, ensure_ascii=False, indent=2)


def assess_decision_risks(sku: str, decision_context: str = "") -> str:
    """从库存、现金流、广告效率、数据质量四个角度审查决策风险。"""
    snapshot = _snapshot_for_sku(sku)
    risks = []
    if snapshot.days_cover < snapshot.lead_time_days:
        risks.append("当前库存覆盖天数小于采购周期，存在断货风险。")
    if snapshot.acos_7d and snapshot.acos_7d > 35:
        risks.append("近 7 日 ACOS 偏高，放大补货前应确认广告转化质量。")
    if snapshot.unit_cost <= 0 or snapshot.selling_price <= 0:
        risks.append("成本或售价数据缺失，利润判断不可靠。")
    if snapshot.daily_sales_7d <= 0 and snapshot.daily_sales_30d <= 0:
        risks.append("销量数据缺失，补货数量只能作为粗略估算。")
    if "涨价" in decision_context or "提价" in decision_context:
        risks.append("价格调整可能影响转化率，建议做小流量测试后再全量执行。")
    if not risks:
        risks.append("未发现明显红线，但仍需人工确认供应链、平台政策和现金流。")
    return json.dumps({"sku": snapshot.to_dict(), "risks": risks}, ensure_ascii=False, indent=2)


def save_decision_report(title: str, content: str) -> str:
    """把最终决策报告保存到 D:\\A2A\\data\\reports 目录。"""
    _ensure_data_dir()
    slug = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff_-]+", "-", title).strip("-")[:60] or "decision-report"
    filename = f"{datetime.now().strftime('%Y%m%d-%H%M%S')}-{slug}.md"
    path = _safe_path(REPORT_DIR / filename)
    evidence_chain = _decision_evidence_chain(report_path=str(path))
    final_content = content
    if "## Evidence Chain" not in final_content:
        final_content = f"{final_content.rstrip()}\n\n{_format_evidence_chain_markdown(evidence_chain)}\n"
    path.write_text(final_content, encoding="utf-8")
    return json.dumps({"saved_to": str(path), "evidence_chain": evidence_chain}, ensure_ascii=False, indent=2)
