from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
from datetime import datetime, timezone
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse

from src.a2a_ecommerce_demo.state_io import atomic_write_json, atomic_write_text, load_json

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = Path(os.getenv("A2A_DATA_DIR", PROJECT_ROOT / "data")).resolve()
RAW_DIR = Path(os.getenv("A2A_RAW_DIR", PROJECT_ROOT / "raw")).resolve()
WAREHOUSE_DIR = Path(os.getenv("A2A_WAREHOUSE_DIR", DATA_DIR / "warehouse")).resolve()
WIKI_DIR = Path(os.getenv("A2A_WIKI_DIR", PROJECT_ROOT / "wiki")).resolve()
TASK_DIR = Path(os.getenv("A2A_TASK_DIR", DATA_DIR / "tasks")).resolve()
AUDIT_DIR = Path(os.getenv("A2A_AUDIT_DIR", DATA_DIR / "audit")).resolve()
SOURCE_REGISTRY_DIR = Path(os.getenv("A2A_SOURCE_REGISTRY_DIR", DATA_DIR / "source_registry")).resolve()
SOURCE_REGISTRY_PATH = Path(os.getenv("A2A_SOURCE_REGISTRY_PATH", SOURCE_REGISTRY_DIR / "sources.json")).resolve()
SNAPSHOT_MANIFEST_PATH = Path(
    os.getenv("A2A_SOURCE_SNAPSHOT_MANIFEST", SOURCE_REGISTRY_DIR / "snapshots.jsonl")
).resolve()
DATASET_REGISTRY_PATH = Path(os.getenv("A2A_DATASET_REGISTRY", WAREHOUSE_DIR / "dataset_registry.json")).resolve()

SOURCE_REGISTRY_SCHEMA = "a2a_source_registry_v1"
SNAPSHOT_MANIFEST_SCHEMA = "a2a_source_snapshot_v1"
SOURCE_SYNC_WORKFLOW_STEPS = [
    "source_watcher",
    "snapshotter",
    "schema_profiler",
    "quality_gate",
    "fact_registrar",
    "wiki_ingest",
    "lightrag_sync",
    "verifier",
]
SUPPORTED_SOURCE_TYPES = [
    "wecom_wedrive_file",
    "wecom_wedrive_folder",
    "local_file",
    "local_folder",
    "manual_upload",
    "wecom_smartsheet",
    "erp_readonly_snapshot",
    "api_pull",
    "mcp_readonly_tool",
    "agent_reach_public_web",
    "agent_reach_public_search",
    "agent_reach_public_video",
    "agent_reach_social",
]
SUPPORTED_SYNC_MODES = ["manual", "on_demand", "polling", "webhook_placeholder"]
SUPPORTED_SOURCE_STATUSES = ["active", "paused", "failed", "disabled", "archived"]
LOCAL_SOURCE_TYPES = {"local_file", "local_folder", "manual_upload"}
SENSITIVE_QUERY_KEYS = {"access_token", "token", "secret", "key", "apikey", "api_key", "scode", "corpsecret"}
SECRET_VALUE_PATTERN = re.compile(r"(sk-|tp-|secret|token|apikey|api_key|password|access_token)", re.IGNORECASE)
ENV_KEY_PATTERN = re.compile(r"^[A-Z][A-Z0-9_]*$")
STRUCTURED_SUFFIXES = {".csv", ".tsv", ".xlsx", ".xlsm", ".json"}


def _now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _compact_now() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def _json(data: dict[str, Any]) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2)


def _ensure_dirs() -> None:
    SOURCE_REGISTRY_DIR.mkdir(parents=True, exist_ok=True)
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    WAREHOUSE_DIR.mkdir(parents=True, exist_ok=True)
    WIKI_DIR.mkdir(parents=True, exist_ok=True)
    TASK_DIR.mkdir(parents=True, exist_ok=True)
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)


def _workspace_root() -> Path:
    configured = os.getenv("A2A_WORKSPACE_DIR", "").strip()
    if configured:
        return Path(configured).expanduser().resolve(strict=False)
    return Path(os.path.commonpath([str(DATA_DIR), str(RAW_DIR), str(WIKI_DIR)])).resolve(strict=False)


def _local_source_allowed_roots() -> list[Path]:
    configured = os.getenv("A2A_LOCAL_SOURCE_ALLOWED_ROOTS", "").strip()
    if configured:
        roots = [Path(item).expanduser().resolve(strict=False) for item in configured.split(os.pathsep) if item.strip()]
    else:
        roots = [_workspace_root(), RAW_DIR]
    deduped: list[Path] = []
    for root in roots:
        if root not in deduped:
            deduped.append(root)
    return deduped


def _error(error_code: str, message: str, **metadata: Any) -> str:
    return _json({"status": "error", "error_code": error_code, "error": message, **metadata})


def _slugify(value: str, fallback: str = "source") -> str:
    slug = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff_-]+", "_", str(value or "")).strip("_")
    return slug[:80] or fallback


def _source_registry_default() -> dict[str, Any]:
    return {
        "schema": SOURCE_REGISTRY_SCHEMA,
        "updated_at": "",
        "registry_path": str(SOURCE_REGISTRY_PATH),
        "supported_source_types": SUPPORTED_SOURCE_TYPES,
        "supported_sync_modes": SUPPORTED_SYNC_MODES,
        "workflow_steps": SOURCE_SYNC_WORKFLOW_STEPS,
        "sources": {},
    }


def _load_registry() -> dict[str, Any]:
    _ensure_dirs()
    registry = load_json(SOURCE_REGISTRY_PATH, _source_registry_default())
    if registry.get("schema") != SOURCE_REGISTRY_SCHEMA:
        registry["schema"] = SOURCE_REGISTRY_SCHEMA
    registry.setdefault("registry_path", str(SOURCE_REGISTRY_PATH))
    registry.setdefault("supported_source_types", SUPPORTED_SOURCE_TYPES)
    registry.setdefault("supported_sync_modes", SUPPORTED_SYNC_MODES)
    registry.setdefault("workflow_steps", SOURCE_SYNC_WORKFLOW_STEPS)
    registry.setdefault("sources", {})
    if not isinstance(registry["sources"], dict):
        registry["sources"] = {}
    return registry


def _save_registry(registry: dict[str, Any]) -> None:
    registry["schema"] = SOURCE_REGISTRY_SCHEMA
    registry["updated_at"] = _now()
    registry["registry_path"] = str(SOURCE_REGISTRY_PATH)
    registry["supported_source_types"] = SUPPORTED_SOURCE_TYPES
    registry["supported_sync_modes"] = SUPPORTED_SYNC_MODES
    registry["workflow_steps"] = SOURCE_SYNC_WORKFLOW_STEPS
    atomic_write_json(SOURCE_REGISTRY_PATH, registry)


def _normalize_credential_env_keys(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        raw = re.split(r"[,;\s]+", value)
    elif isinstance(value, list | tuple):
        raw = [str(item) for item in value]
    else:
        raw = [str(value)]
    keys: list[str] = []
    seen: set[str] = set()
    for item in raw:
        key = str(item or "").strip()
        if not key or key in seen:
            continue
        seen.add(key)
        keys.append(key)
    return keys


def _credential_error(keys: list[str]) -> str:
    for key in keys:
        if not ENV_KEY_PATTERN.match(key) or key.startswith(("sk-", "tp-")):
            return f"credential_env_keys must contain env var names only; got unsafe value {key!r}"
    return ""


def _uri_has_sensitive_query(uri: str) -> bool:
    parsed = urlparse(uri or "")
    query = parse_qs(parsed.query)
    return any(key.lower() in SENSITIVE_QUERY_KEYS for key in query)


def _redact_url(uri: str) -> str:
    parsed = urlparse(uri or "")
    if not parsed.scheme or not parsed.netloc:
        return uri
    query_pairs: list[tuple[str, str]] = []
    for key, values in parse_qs(parsed.query, keep_blank_values=True).items():
        if key.lower() in SENSITIVE_QUERY_KEYS:
            query_pairs.append((key, "***REDACTED***"))
        elif key in {"tab", "sheet", "sheet_id", "viewId"}:
            for value in values:
                query_pairs.append((key, value))
    query = urlencode(query_pairs) if query_pairs else ("..." if parsed.query else "")
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", query, ""))


def _docid_from_smartsheet_url(uri: str) -> str:
    path = urlparse(uri or "").path.strip("/")
    return path.split("/")[-1] if path else ""


def _sheet_id_from_smartsheet_url(uri: str) -> str:
    query = parse_qs(urlparse(uri or "").query)
    return (query.get("tab") or query.get("sheet_id") or [""])[0]


def _sanitize_metadata(source_type: str, metadata: dict[str, Any] | None, uri: str) -> dict[str, Any]:
    metadata = dict(metadata or {})
    sanitized: dict[str, Any] = {}
    if source_type in {"wecom_wedrive_file", "wecom_wedrive_folder"}:
        for key in ["space_id", "file_id", "folder_id", "file_name", "path_summary", "remote_hash", "mtime", "size"]:
            if metadata.get(key) not in (None, ""):
                sanitized[key] = metadata[key]
        if not sanitized.get("space_id") and metadata.get("spaceid"):
            sanitized["space_id"] = metadata["spaceid"]
        if not sanitized.get("file_id") and metadata.get("fileid"):
            sanitized["file_id"] = metadata["fileid"]
        if not sanitized.get("file_name") and metadata.get("name"):
            sanitized["file_name"] = metadata["name"]
        return sanitized
    if source_type == "wecom_smartsheet":
        doc_url = str(metadata.get("doc_url") or uri or "")
        docid = str(metadata.get("docid") or _docid_from_smartsheet_url(doc_url)).strip()
        sheet_ids = metadata.get("sheet_ids")
        if isinstance(sheet_ids, str):
            sheet_ids = [item for item in re.split(r"[,;\s]+", sheet_ids) if item]
        if not sheet_ids:
            sheet_id = str(metadata.get("sheet_id") or _sheet_id_from_smartsheet_url(doc_url)).strip()
            sheet_ids = [sheet_id] if sheet_id else []
        sanitized.update(
            {
                "docid": docid,
                "sheet_ids": [str(item) for item in sheet_ids],
                "dataset": str(metadata.get("dataset") or "smart_records"),
                "doc_url_summary": _redact_url(doc_url),
            }
        )
        return sanitized
    if source_type == "erp_readonly_snapshot":
        for key in ["connector_id", "dataset", "filters", "query_filters", "live_read_only_fallback"]:
            if key in metadata:
                sanitized[key] = metadata[key]
        return sanitized
    for key, value in metadata.items():
        lower = str(key).lower()
        if lower in {"download_url", "temporary_url", "access_token", "token", "secret", "apikey", "api_key"}:
            continue
        sanitized[key] = _redact_url(str(value)) if "url" in lower and isinstance(value, str) else value
    return sanitized


def _default_uri_for(source_type: str, metadata: dict[str, Any], uri: str) -> str:
    if uri:
        return _redact_url(uri) if source_type == "wecom_smartsheet" else uri
    if source_type in {"wecom_wedrive_file", "wecom_wedrive_folder"}:
        space_id = str(metadata.get("space_id") or "")
        object_id = str(metadata.get("file_id") or metadata.get("folder_id") or "")
        kind = "folder" if source_type == "wecom_wedrive_folder" else "file"
        return f"wecom-wedrive://{space_id}/{kind}/{object_id}".rstrip("/")
    if source_type == "wecom_smartsheet":
        docid = str(metadata.get("docid") or "")
        return f"wecom-smartsheet://{docid}".rstrip("/")
    if source_type == "erp_readonly_snapshot":
        connector_id = str(metadata.get("connector_id") or "")
        dataset = str(metadata.get("dataset") or "")
        return f"erp-readonly://{connector_id}/{dataset}".rstrip("/")
    return ""


def _path_is_under(path: Path, root: Path) -> bool:
    resolved = path.expanduser().resolve(strict=False)
    root_resolved = root.expanduser().resolve(strict=False)
    return root_resolved in [resolved, *resolved.parents]


def _path_is_under_any_allowed_root(path: Path) -> bool:
    return any(_path_is_under(path, root) for root in _local_source_allowed_roots())


def _validate_local_path(source_type: str, uri: str, allowed_root: str) -> tuple[str, str]:
    if source_type not in LOCAL_SOURCE_TYPES:
        return uri, allowed_root
    if not uri:
        raise ValueError("local/manual source requires uri")
    root = Path(allowed_root or RAW_DIR).expanduser().resolve(strict=False)
    target = Path(uri).expanduser().resolve(strict=False)
    if not _path_is_under_any_allowed_root(root):
        raise PermissionError("allowed_root_outside_workspace")
    if not _path_is_under(target, root):
        raise PermissionError("path_outside_allowed_root")
    return str(target), str(root)


def _source_summary(source: dict[str, Any], snapshots: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    snapshots = snapshots or []
    source_snapshots = [item for item in snapshots if item.get("source_id") == source.get("source_id")]
    latest = source_snapshots[-1] if source_snapshots else {}
    return {
        **source,
        "latest_snapshot": latest,
        "snapshot_count": len(source_snapshots),
        "last_snapshot_at": latest.get("observed_at", ""),
        "last_row_count": latest.get("row_count", 0),
        "last_schema_hash": latest.get("schema_hash", ""),
        "freshness_status": _freshness_status(source, latest),
    }


def _freshness_status(source: dict[str, Any], latest: dict[str, Any]) -> str:
    if source.get("status") == "failed" or latest.get("status") == "failed":
        return "failed"
    if not latest:
        return "never_synced"
    return "fresh"


def _freshness_summary(sources: dict[str, dict[str, Any]], snapshots: list[dict[str, Any]]) -> dict[str, Any]:
    latest_by_source: dict[str, dict[str, Any]] = {}
    for snapshot in snapshots:
        latest_by_source[str(snapshot.get("source_id") or "")] = snapshot
    failed = 0
    recent_success = 0
    stale = 0
    for source_id, source in sources.items():
        latest = latest_by_source.get(source_id, {})
        status = _freshness_status(source, latest)
        if status == "failed":
            failed += 1
        elif status == "fresh":
            recent_success += 1
        else:
            stale += 1
    return {
        "failed_count": failed,
        "recent_success_count": recent_success,
        "stale_count": stale,
        "total": len(sources),
    }


def register_source(
    source_id: str,
    display_name: str,
    source_type: str,
    uri: str = "",
    allowed_root: str = "",
    sync_mode: str = "on_demand",
    owner: str = "",
    sensitivity_level: str = "internal",
    credential_env_keys: list[str] | str | None = None,
    format_hint: str = "",
    expected_schema: dict[str, Any] | list[str] | None = None,
    freshness_sla: str = "",
    status: str = "active",
    metadata: dict[str, Any] | None = None,
) -> str:
    """Register a long-lived read source without persisting secrets or temporary download URLs."""
    try:
        registry = _load_registry()
        final_source_id = _slugify(source_id, "source")
        if final_source_id in registry["sources"]:
            return _error("duplicate_source_id", f"Source already exists: {final_source_id}")
        if source_type not in SUPPORTED_SOURCE_TYPES:
            return _error("unsupported_source_type", f"Unsupported source_type: {source_type}")
        if sync_mode not in SUPPORTED_SYNC_MODES:
            return _error("unsupported_sync_mode", f"Unsupported sync_mode: {sync_mode}")
        if status not in SUPPORTED_SOURCE_STATUSES:
            return _error("unsupported_source_status", f"Unsupported source status: {status}")
        keys = _normalize_credential_env_keys(credential_env_keys)
        credential_error = _credential_error(keys)
        if credential_error:
            return _error("credential_plaintext", credential_error)
        if source_type not in {"wecom_smartsheet"} and _uri_has_sensitive_query(uri):
            return _error("unsafe_source_uri", "Source uri must not contain tokens, scode, apikey, or secrets.")
        sanitized_metadata = _sanitize_metadata(source_type, metadata, uri)
        final_uri = _default_uri_for(source_type, sanitized_metadata, uri)
        try:
            final_uri, final_allowed_root = _validate_local_path(source_type, final_uri, allowed_root)
        except PermissionError:
            return _error("path_outside_allowed_root", "Local source path is outside allowed_root.")
        source = {
            "source_id": final_source_id,
            "display_name": display_name or final_source_id,
            "source_type": source_type,
            "uri": final_uri,
            "allowed_root": final_allowed_root if source_type in LOCAL_SOURCE_TYPES else str(allowed_root or ""),
            "sync_mode": sync_mode,
            "owner": owner,
            "sensitivity_level": sensitivity_level or "internal",
            "credential_env_keys": keys,
            "format_hint": format_hint,
            "expected_schema": expected_schema or {},
            "freshness_sla": freshness_sla,
            "status": status,
            "read_only": True,
            "external_write_enabled": False,
            "governance": {
                "scope": "read_only",
                "external_write_enabled": False,
                "source_system_writes_in_scope": False,
            },
            "metadata": sanitized_metadata,
            "created_at": _now(),
            "updated_at": _now(),
        }
        registry["sources"][final_source_id] = source
        _save_registry(registry)
        return _json({"status": "registered", "source": source, "registry_path": str(SOURCE_REGISTRY_PATH)})
    except Exception as exc:
        return _error("register_source_failed", str(exc))


def list_sources() -> str:
    """List registered long-lived sources with latest snapshot and freshness summary."""
    registry = _load_registry()
    snapshots = load_snapshots()
    sources = {
        source_id: _source_summary(source, snapshots)
        for source_id, source in sorted(registry.get("sources", {}).items())
    }
    return _json(
        {
            **registry,
            "sources": sources,
            "source_count": len(sources),
            "freshness": _freshness_summary(registry.get("sources", {}), snapshots),
        }
    )


def get_source(source_id: str) -> str:
    """Return one source and its latest snapshot metadata."""
    registry = _load_registry()
    source = registry.get("sources", {}).get(source_id)
    if not isinstance(source, dict):
        return _error("source_not_found", f"Source not found: {source_id}")
    return _json({"status": "success", "source": _source_summary(source, load_snapshots())})


def set_source_status(source_id: str, status: str) -> str:
    """Pause, resume, disable, archive, or mark a source failed without touching the external system."""
    registry = _load_registry()
    source = registry.get("sources", {}).get(source_id)
    if not isinstance(source, dict):
        return _error("source_not_found", f"Source not found: {source_id}")
    if status not in SUPPORTED_SOURCE_STATUSES:
        return _error("unsupported_source_status", f"Unsupported source status: {status}")
    source["status"] = status
    source["updated_at"] = _now()
    _save_registry(registry)
    return _json({"status": "updated", "source": source})


def rebind_source_path(
    source_id: str,
    uri: str = "",
    allowed_root: str = "",
    metadata: dict[str, Any] | None = None,
) -> str:
    """Update a source location while preserving the same source_id and governance guardrails."""
    registry = _load_registry()
    source = registry.get("sources", {}).get(source_id)
    if not isinstance(source, dict):
        return _error("source_not_found", f"Source not found: {source_id}")
    source_type = str(source.get("source_type") or "")
    if source_type not in {"wecom_smartsheet"} and _uri_has_sensitive_query(uri):
        return _error("unsafe_source_uri", "Source uri must not contain tokens, scode, apikey, or secrets.")
    merged_metadata = {**dict(source.get("metadata") or {}), **dict(metadata or {})}
    sanitized_metadata = _sanitize_metadata(source_type, merged_metadata, uri or str(source.get("uri") or ""))
    final_uri = _default_uri_for(source_type, sanitized_metadata, uri or str(source.get("uri") or ""))
    try:
        final_uri, final_allowed_root = _validate_local_path(
            source_type,
            final_uri,
            allowed_root or str(source.get("allowed_root") or ""),
        )
    except PermissionError:
        return _error("path_outside_allowed_root", "Local source path is outside allowed_root.")
    source["uri"] = final_uri
    source["allowed_root"] = final_allowed_root if source_type in LOCAL_SOURCE_TYPES else str(allowed_root or source.get("allowed_root") or "")
    source["metadata"] = sanitized_metadata
    source["updated_at"] = _now()
    _save_registry(registry)
    return _json({"status": "updated", "source": source})


def load_snapshots() -> list[dict[str, Any]]:
    if not SNAPSHOT_MANIFEST_PATH.exists():
        return []
    records: list[dict[str, Any]] = []
    for line in SNAPSHOT_MANIFEST_PATH.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            value = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(value, dict):
            records.append(value)
    return records


def _append_snapshot(snapshot: dict[str, Any]) -> None:
    _ensure_dirs()
    SNAPSHOT_MANIFEST_PATH.parent.mkdir(parents=True, exist_ok=True)
    with SNAPSHOT_MANIFEST_PATH.open("a", encoding="utf-8") as file:
        file.write(json.dumps(snapshot, ensure_ascii=False, sort_keys=True) + "\n")


def list_source_snapshots(source_id: str = "", limit: int = 200) -> str:
    """List immutable raw snapshots, optionally scoped to one source."""
    snapshots = load_snapshots()
    if source_id:
        snapshots = [item for item in snapshots if item.get("source_id") == source_id]
    return _json(
        {
            "schema": SNAPSHOT_MANIFEST_SCHEMA,
            "manifest_path": str(SNAPSHOT_MANIFEST_PATH),
            "source_id": source_id,
            "snapshot_count": len(snapshots),
            "snapshots": snapshots[-max(1, min(int(limit or 200), 1000)) :],
        }
    )


def _safe_file_name(value: str, fallback: str = "original.csv") -> str:
    name = Path(str(value or fallback)).name
    name = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff_.-]+", "_", name).strip("._")
    return name or fallback


def _sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _profile_csv(content: bytes, file_name: str) -> dict[str, Any]:
    text = _decode_text(content)
    sample = text[:4096]
    delimiter = "\t" if Path(file_name).suffix.lower() == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        delimiter = dialect.delimiter
    except csv.Error:
        pass
    reader = csv.reader(StringIO(text), delimiter=delimiter)
    rows = list(reader)
    headers = [str(item).strip() for item in rows[0]] if rows else []
    row_count = max(0, len(rows) - 1)
    sheet_name = Path(file_name).stem or "sheet1"
    return {
        "row_count": row_count,
        "sheet_names": [sheet_name],
        "schemas": {sheet_name: headers},
        "field_profiles": [{"field": header} for header in headers if header],
    }


def _profile_json(content: bytes, file_name: str) -> dict[str, Any]:
    try:
        payload = json.loads(_decode_text(content))
    except json.JSONDecodeError:
        payload = {}
    rows = payload.get("rows") if isinstance(payload, dict) else payload if isinstance(payload, list) else []
    if not isinstance(rows, list):
        rows = []
    headers: list[str] = []
    for row in rows:
        if isinstance(row, dict):
            for key in row:
                if key not in headers:
                    headers.append(str(key))
    sheet_name = Path(file_name).stem or "json"
    return {
        "row_count": len(rows),
        "sheet_names": [sheet_name],
        "schemas": {sheet_name: headers},
        "field_profiles": [{"field": header} for header in headers],
    }


def _profile_excel(content: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "row_count": 0,
            "sheet_names": [],
            "schemas": {},
            "field_profiles": [],
            "warnings": ["openpyxl_missing"],
        }
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    schemas: dict[str, list[str]] = {}
    total_rows = 0
    fields: list[dict[str, str]] = []
    for sheet in workbook.worksheets:
        first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(value).strip() for value in first or [] if value is not None]
        schemas[sheet.title] = headers
        total_rows += max(0, sheet.max_row - 1)
        fields.extend({"field": header, "sheet": sheet.title} for header in headers)
    return {
        "row_count": total_rows,
        "sheet_names": list(schemas),
        "schemas": schemas,
        "field_profiles": fields,
    }


def _profile_content(content: bytes, file_name: str) -> dict[str, Any]:
    suffix = Path(file_name).suffix.lower()
    if suffix in {".csv", ".tsv"}:
        profile = _profile_csv(content, file_name)
    elif suffix in {".xlsx", ".xlsm"}:
        profile = _profile_excel(content)
    elif suffix == ".json":
        profile = _profile_json(content, file_name)
    else:
        profile = {"row_count": 0, "sheet_names": [], "schemas": {}, "field_profiles": []}
    schema_hash = hashlib.sha256(json.dumps(profile.get("schemas", {}), ensure_ascii=False, sort_keys=True).encode()).hexdigest()
    profile["schema_hash"] = schema_hash
    return profile


def _latest_success_snapshot(source_id: str) -> dict[str, Any] | None:
    for snapshot in reversed(load_snapshots()):
        if snapshot.get("source_id") == source_id and snapshot.get("status") == "success":
            return snapshot
    return None


def _existing_snapshot_by_sha(source_id: str, sha256: str) -> dict[str, Any] | None:
    for snapshot in reversed(load_snapshots()):
        if snapshot.get("source_id") == source_id and snapshot.get("sha256") == sha256 and snapshot.get("status") == "success":
            return snapshot
    return None


def _relative(path: Path) -> str:
    resolved = Path(path).resolve(strict=False)
    for prefix, root in [("raw", RAW_DIR), ("data", DATA_DIR), ("wiki", WIKI_DIR)]:
        try:
            return f"{prefix}/{resolved.relative_to(root).as_posix()}"
        except ValueError:
            continue
    return str(resolved)


def _dataset_slug(source_id: str, snapshot_id: str) -> str:
    return _slugify(f"{source_id}_{snapshot_id}", "source_snapshot")[:120]


def _load_dataset_registry() -> dict[str, Any]:
    registry = load_json(DATASET_REGISTRY_PATH, {"schema": "a2a_dataset_registry_v1", "updated_at": "", "datasets": {}})
    registry.setdefault("schema", "a2a_dataset_registry_v1")
    registry.setdefault("datasets", {})
    if not isinstance(registry["datasets"], dict):
        registry["datasets"] = {}
    return registry


def _write_dataset_registry_entry(snapshot: dict[str, Any]) -> None:
    registry = _load_dataset_registry()
    slug = snapshot["duckdb_dataset_slug"]
    registry["datasets"][slug] = {
        "dataset_slug": slug,
        "source": snapshot["raw_snapshot_path"],
        "relative_source": _relative(Path(snapshot["raw_snapshot_path"])),
        "source_id": snapshot["source_id"],
        "snapshot_id": snapshot["snapshot_id"],
        "source_type": snapshot["source_type"],
        "source_snapshot_path": snapshot["raw_snapshot_path"],
        "registered_at": snapshot["observed_at"],
        "row_count": snapshot["row_count"],
        "schema_hash": snapshot["schema_hash"],
        "field_profiles": snapshot.get("field_profiles", []),
        "sheet_views": [
            {
                "sheet": sheet,
                "headers": snapshot.get("schema", {}).get(sheet, []),
                "rows": snapshot["row_count"] if index == 0 else 0,
            }
            for index, sheet in enumerate(snapshot.get("sheet_names", []))
        ],
        "source_registry": str(SOURCE_REGISTRY_PATH),
        "source_snapshot_manifest": str(SNAPSHOT_MANIFEST_PATH),
        "wiki_pages": snapshot.get("wiki_pages", []),
        "derived_exports": [],
        "mart_views": [],
        "semantic_views": [],
    }
    registry["updated_at"] = _now()
    atomic_write_json(DATASET_REGISTRY_PATH, registry)


def _write_source_wiki_page(source: dict[str, Any], snapshot: dict[str, Any]) -> str:
    path = WIKI_DIR / "sources" / f"{source['source_id']}.md"
    lines = [
        "---",
        "type: source",
        f"source_id: {source['source_id']}",
        f"source_type: {source['source_type']}",
        f"updated_at: {snapshot['observed_at']}",
        "evidence:",
        f"  - {snapshot['raw_snapshot_path']}",
        "---",
        "",
        f"# {source.get('display_name') or source['source_id']}",
        "",
        f"- Latest snapshot: `{snapshot['snapshot_id']}`",
        f"- Dataset slug: `{snapshot['duckdb_dataset_slug']}`",
        f"- Row count: {snapshot.get('row_count', 0)}",
        f"- Schema hash: `{snapshot.get('schema_hash', '')}`",
    ]
    atomic_write_text(path, "\n".join(lines).rstrip() + "\n")
    return str(path)


def _append_wiki_log(source: dict[str, Any], snapshot: dict[str, Any]) -> None:
    path = WIKI_DIR / "log.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    existing = path.read_text(encoding="utf-8") if path.exists() else "# Wiki Log\n"
    entry = (
        f"\n## [{datetime.now().strftime('%Y-%m-%d')}] Source snapshot synced\n\n"
        f"- source_id: `{source['source_id']}`\n"
        f"- snapshot_id: `{snapshot['snapshot_id']}`\n"
        f"- raw_snapshot_path: `{snapshot['raw_snapshot_path']}`\n"
        f"- dataset: `{snapshot['duckdb_dataset_slug']}`\n"
    )
    atomic_write_text(path, existing.rstrip() + "\n" + entry)


def _redact_text(value: str) -> str:
    text = str(value)
    text = re.sub(r"(?i)(access_token|apikey|api_key|scode|token|secret|password)=([^&\s\"',}]+)", r"\1=***REDACTED***", text)
    text = re.sub(r"(?i)(sk-|tp-)[A-Za-z0-9_-]{8,}", "***REDACTED***", text)
    return text


def _redact_value(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: _redact_value(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_redact_value(item) for item in value]
    if isinstance(value, str):
        return _redact_text(value)
    return value


def _record_source_audit(event_type: str, summary: str, source_id: str, status: str, metadata: dict[str, Any]) -> str:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    event_id = f"audit_{datetime.now().strftime('%Y%m%d%H%M%S')}_{hashlib.sha256((event_type + source_id + status).encode()).hexdigest()[:8]}"
    event = {
        "event_id": event_id,
        "timestamp": _now(),
        "created_at": _now(),
        "event_type": event_type,
        "actor": metadata.get("requested_by", "agent"),
        "tool_name": metadata.get("tool_name", "source_registry_tools"),
        "data_sources": ["source_registry", source_id],
        "status": status,
        "summary": _redact_text(summary),
        "metadata": _redact_value(metadata),
    }
    with (AUDIT_DIR / "events.jsonl").open("a", encoding="utf-8") as file:
        file.write(json.dumps(event, ensure_ascii=False) + "\n")
    return event_id


def _update_source_after_sync(source_id: str, update: dict[str, Any]) -> None:
    registry = _load_registry()
    source = registry.get("sources", {}).get(source_id)
    if isinstance(source, dict):
        source.update(update)
        source["updated_at"] = _now()
        _save_registry(registry)


def _build_snapshot_record(
    source: dict[str, Any],
    *,
    file_name: str,
    content: bytes,
    source_mtime: str,
    source_size: int,
    profile_extra: dict[str, Any] | None = None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    sha256 = _sha256_bytes(content)
    existing = _existing_snapshot_by_sha(source["source_id"], sha256)
    profile = _profile_content(content, file_name)
    if existing:
        return existing, {"changed": False, "quality_warnings": []}
    previous = _latest_success_snapshot(source["source_id"])
    quality_warnings: list[str] = []
    if previous and previous.get("schema_hash") and previous.get("schema_hash") != profile.get("schema_hash"):
        quality_warnings.append("schema_hash_changed: source schema drift requires quality review")
    snapshot_dir = RAW_DIR / "snapshots" / source["source_id"] / f"{_compact_now()}-{sha256[:10]}"
    snapshot_dir.mkdir(parents=True, exist_ok=False)
    original_name = _safe_file_name(file_name, "original.csv")
    target = snapshot_dir / f"original{Path(original_name).suffix or '.bin'}"
    target.write_bytes(content)
    snapshot_id = snapshot_dir.name
    snapshot = {
        "record_schema": SNAPSHOT_MANIFEST_SCHEMA,
        "snapshot_id": snapshot_id,
        "source_id": source["source_id"],
        "source_type": source["source_type"],
        "observed_at": _now(),
        "source_mtime": source_mtime,
        "source_size": int(source_size or len(content)),
        "sha256": sha256,
        "remote_hash": str((profile_extra or {}).get("remote_hash") or ""),
        "schema_hash": profile["schema_hash"],
        "schema": profile.get("schemas", {}),
        "row_count": int(profile.get("row_count", 0) or 0),
        "sheet_names": profile.get("sheet_names", []),
        "raw_snapshot_path": str(target),
        "cleaned_paths": [],
        "duckdb_dataset_slug": _dataset_slug(source["source_id"], snapshot_id),
        "wiki_pages": [],
        "lightrag_docs": [],
        "task_id": str((profile_extra or {}).get("task_id") or ""),
        "audit_event_id": "",
        "status": "success",
        "profile": {**profile, **(profile_extra or {}).get("profile", {})},
        "quality_warnings": quality_warnings,
        "field_profiles": profile.get("field_profiles", []),
    }
    return snapshot, {"changed": True, "quality_warnings": quality_warnings}


def _finalize_snapshot(source: dict[str, Any], snapshot: dict[str, Any], requested_by: str) -> dict[str, Any]:
    wiki_page = _write_source_wiki_page(source, snapshot)
    snapshot["wiki_pages"] = [wiki_page]
    _write_dataset_registry_entry(snapshot)
    _append_wiki_log(source, snapshot)
    audit_event_id = _record_source_audit(
        "source_sync_completed",
        f"Source snapshot synced: {source['source_id']}/{snapshot['snapshot_id']}",
        source["source_id"],
        "success",
        {"snapshot_id": snapshot["snapshot_id"], "requested_by": requested_by, "raw_snapshot_path": snapshot["raw_snapshot_path"]},
    )
    snapshot["audit_event_id"] = audit_event_id
    _append_snapshot(snapshot)
    _update_source_after_sync(
        source["source_id"],
        {
            "last_sync": {
                "status": "success",
                "snapshot_id": snapshot["snapshot_id"],
                "snapshot_at": snapshot["observed_at"],
                "row_count": snapshot["row_count"],
                "schema_hash": snapshot["schema_hash"],
            }
        },
    )
    return snapshot


def _local_file_payload(source: dict[str, Any]) -> dict[str, Any]:
    try:
        validated_path, validated_root = _validate_local_path(
            str(source.get("source_type") or "local_file"),
            str(source.get("uri") or ""),
            str(source.get("allowed_root") or RAW_DIR),
        )
    except ValueError as exc:
        raise PermissionError(str(exc)) from exc
    path = Path(validated_path)
    allowed_root = Path(validated_root)
    if not _path_is_under(path, allowed_root):
        raise PermissionError("Local source path is outside allowed_root")
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Local source file is missing: {path}")
    stat = path.stat()
    return {
        "file_name": path.name,
        "content": path.read_bytes(),
        "source_mtime": datetime.fromtimestamp(stat.st_mtime, timezone.utc).replace(microsecond=0).isoformat(),
        "source_size": stat.st_size,
        "profile": {"adapter": source["source_type"]},
    }


def _local_folder_payload(source: dict[str, Any]) -> dict[str, Any]:
    try:
        validated_folder, validated_root = _validate_local_path(
            "local_folder",
            str(source.get("uri") or ""),
            str(source.get("allowed_root") or source.get("uri") or RAW_DIR),
        )
    except ValueError as exc:
        raise PermissionError(str(exc)) from exc
    folder = Path(validated_folder)
    allowed_root = Path(validated_root)
    if not _path_is_under(folder, allowed_root):
        raise PermissionError("Local source folder is outside allowed_root")
    candidates = [
        path
        for path in folder.rglob("*")
        if path.is_file() and path.suffix.lower() in STRUCTURED_SUFFIXES and _path_is_under(path, allowed_root)
    ]
    if not candidates:
        raise FileNotFoundError(f"No structured files found under local source folder: {folder}")
    latest = max(candidates, key=lambda item: item.stat().st_mtime)
    return _local_file_payload({**source, "source_type": "local_file", "uri": str(latest), "allowed_root": str(allowed_root)})


def _fetch_wecom_wedrive_source(source: dict[str, Any]) -> dict[str, Any]:
    """Read-only seam for enterprise WeCom Wedrive.

    Real API credentials are intentionally not required here. Tests or local operators can
    provide `metadata.local_fixture_path`; production can replace this seam with an official
    list/download client that still returns this same shape.
    """
    fixture = str(source.get("metadata", {}).get("local_fixture_path") or "").strip()
    if fixture:
        path = Path(fixture).expanduser().resolve(strict=False)
        if not path.exists() or not path.is_file():
            raise FileNotFoundError(f"Wedrive fixture file is missing: {path}")
        stat = path.stat()
        return {
            "file_name": source.get("metadata", {}).get("file_name") or path.name,
            "content": path.read_bytes(),
            "source_mtime": datetime.fromtimestamp(stat.st_mtime, timezone.utc).replace(microsecond=0).isoformat(),
            "source_size": stat.st_size,
            "profile": {"adapter": "wecom_wedrive", "space_id": source.get("metadata", {}).get("space_id", "")},
        }
    raise RuntimeError(
        "wecom_wedrive_client_not_configured: fill WeCom Wedrive app credentials in env and wire the read-only client, "
        "or provide metadata.local_fixture_path for local verification."
    )


def _rows_to_csv_bytes(rows: list[dict[str, Any]]) -> bytes:
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(str(key))
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=headers or ["value"])
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in headers})
    return output.getvalue().encode("utf-8")


def _fetch_wecom_smartsheet_source(source: dict[str, Any]) -> dict[str, Any]:
    from src.a2a_ecommerce_demo.wecom_smartsheet_tools import query_wecom_smartsheet_records

    metadata = source.get("metadata", {})
    payload = json.loads(
        query_wecom_smartsheet_records(
            docid=str(metadata.get("docid") or ""),
            sheet_ids=",".join(str(item) for item in metadata.get("sheet_ids", [])),
            dataset=str(metadata.get("dataset") or "smart_records"),
            limit=5000,
            requested_by="source_registry",
        )
    )
    if payload.get("status") != "success":
        raise RuntimeError(payload.get("error") or "WeCom smart sheet read failed")
    rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
    return {
        "file_name": f"{source['source_id']}.csv",
        "content": _rows_to_csv_bytes(rows),
        "source_mtime": payload.get("queried_at") or _now(),
        "source_size": len(rows),
        "profile": {
            "transport": payload.get("transport", "wecom_wedoc_mcp"),
            "row_count": payload.get("row_count", len(rows)),
            "live_read_only_mcp": True,
        },
    }


def _fetch_erp_readonly_source(source: dict[str, Any]) -> dict[str, Any]:
    from src.a2a_ecommerce_demo.connector_live_tools import query_erp_live_snapshot

    metadata = source.get("metadata", {})
    connector_id = str(metadata.get("connector_id") or "")
    dataset = str(metadata.get("dataset") or "")
    if not connector_id or not dataset:
        raise ValueError("erp_readonly_snapshot source requires metadata.connector_id and metadata.dataset")
    payload = json.loads(query_erp_live_snapshot(connector_id, dataset, limit=1000))
    if payload.get("status") not in {"success", "partial"}:
        raise RuntimeError(payload.get("error") or "ERP readonly snapshot failed")
    return {
        "file_name": f"{connector_id}_{dataset}.json",
        "content": json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        "source_mtime": payload.get("queried_at") or _now(),
        "source_size": len(json.dumps(payload, ensure_ascii=False).encode("utf-8")),
        "profile": {
            "live_read_only_fallback": True,
            "query_filters": metadata.get("filters") or metadata.get("query_filters") or {},
            "row_count": payload.get("row_count", 0),
        },
    }


def _fetch_source_payload(source: dict[str, Any]) -> dict[str, Any]:
    source_type = source.get("source_type")
    if source_type in {"local_file", "manual_upload"}:
        return _local_file_payload(source)
    if source_type == "local_folder":
        return _local_folder_payload(source)
    if source_type in {"wecom_wedrive_file", "wecom_wedrive_folder"}:
        return _fetch_wecom_wedrive_source(source)
    if source_type == "wecom_smartsheet":
        return _fetch_wecom_smartsheet_source(source)
    if source_type == "erp_readonly_snapshot":
        return _fetch_erp_readonly_source(source)
    raise NotImplementedError(f"{source_type} adapter is registered but not enabled in this local build.")


def sync_source(source_id: str, requested_by: str = "agent", task_id: str = "") -> str:
    """Synchronize one source into an immutable raw snapshot if content changed."""
    registry = _load_registry()
    source = registry.get("sources", {}).get(source_id)
    if not isinstance(source, dict):
        return _error("source_not_found", f"Source not found: {source_id}")
    if source.get("status") not in {"active"}:
        return _error("source_not_active", f"Source is not active: {source_id}", source_status=source.get("status"))
    try:
        payload = _fetch_source_payload(source)
        content = payload.get("content", b"")
        if isinstance(content, str):
            content = content.encode("utf-8")
        if not isinstance(content, bytes):
            raise TypeError("adapter content must be bytes or string")
        file_name = _safe_file_name(str(payload.get("file_name") or source_id))
        source_mtime = str(payload.get("source_mtime") or _now())
        source_size = int(payload.get("source_size") or len(content))
        snapshot, meta = _build_snapshot_record(
            source,
            file_name=file_name,
            content=content,
            source_mtime=source_mtime,
            source_size=source_size,
            profile_extra={**payload, "task_id": task_id},
        )
        if not meta["changed"]:
            _record_source_audit(
                "source_sync_skipped",
                f"Source unchanged: {source_id}/{snapshot['snapshot_id']}",
                source_id,
                "skipped_unchanged",
                {"snapshot_id": snapshot["snapshot_id"], "requested_by": requested_by, "tool_name": "sync_source"},
            )
            _update_source_after_sync(
                source_id,
                {
                    "last_sync": {
                        "status": "skipped_unchanged",
                        "snapshot_id": snapshot["snapshot_id"],
                        "snapshot_at": _now(),
                        "row_count": snapshot.get("row_count", 0),
                        "schema_hash": snapshot.get("schema_hash", ""),
                    }
                },
            )
            return _json(
                {
                    "status": "skipped_unchanged",
                    "source_id": source_id,
                    "snapshot_id": snapshot["snapshot_id"],
                    "changed": False,
                    "raw_snapshot_path": snapshot["raw_snapshot_path"],
                    "profile": snapshot.get("profile", {}),
                    "quality_warnings": [],
                    "next_actions": ["No downstream ingest needed; content hash is unchanged."],
                }
            )
        snapshot = _finalize_snapshot(source, snapshot, requested_by)
        return _json(
            {
                "status": "success",
                "source_id": source_id,
                "source_type": source.get("source_type"),
                "snapshot_id": snapshot["snapshot_id"],
                "changed": True,
                "raw_snapshot_path": snapshot["raw_snapshot_path"],
                "snapshot": snapshot,
                "profile": snapshot.get("profile", {}),
                "quality_warnings": snapshot.get("quality_warnings", []),
                "next_actions": [
                    "Review quality warnings before treating schema drift as trusted.",
                    "Use the generated dataset registry entry for DuckDB/wiki/LightRAG downstream work.",
                ],
            }
        )
    except Exception as exc:
        message = _redact_text(str(exc))
        _record_source_audit(
            "source_sync_failed",
            f"Source sync failed: {source_id}: {message}",
            source_id,
            "failed",
            {"requested_by": requested_by, "error": message, "tool_name": "sync_source"},
        )
        _update_source_after_sync(source_id, {"status": "failed", "last_sync": {"status": "failed", "error": message, "snapshot_at": _now()}})
        return _json(
            {
                "status": "failed",
                "error_code": "source_sync_failed",
                "source_id": source_id,
                "changed": False,
                "error": message,
                "quality_warnings": [],
                "next_actions": ["Check source credentials, read permission, file location, or adapter configuration."],
            }
        )


def run_source_sync_workflow(source_id: str, requested_by: str = "agent") -> str:
    """Run the fixed P16 source sync workflow and record task/audit metadata."""
    from src.a2a_ecommerce_demo.task_queue import get_task_queue

    task_id = f"{_compact_now()}-{_slugify(source_id, 'source-sync')}-source-sync"
    task_path = TASK_DIR / f"{task_id}.json"
    queue = get_task_queue()
    durable = queue.enqueue(
        goal=f"source.sync {source_id}",
        requested_by=requested_by,
        idempotency_key=f"source.sync:{source_id}:{_compact_now()}",
        task_id=task_id,
        status="queued",
        payload={"task_type": "source.sync", "source_id": source_id, "workflow_steps": SOURCE_SYNC_WORKFLOW_STEPS},
        json_path=str(task_path),
    )
    steps: list[dict[str, Any]] = []
    for step in SOURCE_SYNC_WORKFLOW_STEPS:
        steps.append({"task": step, "status": "pending", "summary": "", "evidence": [], "data": {}})
    sync_result = json.loads(sync_source(source_id, requested_by=requested_by, task_id=task_id))
    final_status = "success" if sync_result.get("status") in {"success", "skipped_unchanged"} else "failed"
    for step in steps:
        step["status"] = "success" if final_status == "success" else "failed"
        step["summary"] = f"{step['task']} completed for {source_id}." if final_status == "success" else sync_result.get("error", "source sync failed")
        step["completed_at"] = _now()
    if sync_result.get("raw_snapshot_path"):
        steps[1]["evidence"].append(sync_result["raw_snapshot_path"])
    task = {
        "task_id": task_id,
        "task_type": "source.sync",
        "goal": f"source.sync {source_id}",
        "source_id": source_id,
        "requested_by": requested_by,
        "status": final_status,
        "created_at": durable["created_at"],
        "updated_at": _now(),
        "steps": steps,
        "result": sync_result,
        "path": str(task_path),
    }
    atomic_write_json(task_path, task)
    if final_status == "success":
        queue.complete(task_id, result=sync_result)
    else:
        queue.fail(task_id, reason=sync_result.get("error", "source sync failed"), retryable=False)
    return _json({"status": final_status, "task_id": task_id, "task_path": str(task_path), "result": sync_result})


def check_source_registry_health() -> str:
    """Run P16 doctor checks for source metadata, snapshot manifest, and token/path hygiene."""
    problems: list[dict[str, Any]] = []
    registry_payload: dict[str, Any] = {}
    try:
        registry_payload = _load_registry()
    except Exception as exc:
        return _json(
            {
                "status": "fail",
                "problem_count": 1,
                "problems": [{"code": "registry_unreadable", "message": str(exc), "path": str(SOURCE_REGISTRY_PATH)}],
            }
        )
    sources = registry_payload.get("sources", {})
    seen: set[str] = set()
    for source_id, source in sources.items():
        record = source if isinstance(source, dict) else {}
        if source_id in seen:
            problems.append({"code": "duplicate_source_id", "source_id": source_id})
        seen.add(source_id)
        if not record.get("owner"):
            problems.append({"code": "missing_owner", "source_id": source_id})
        if not record.get("freshness_sla"):
            problems.append({"code": "missing_freshness_sla", "source_id": source_id})
        keys = _normalize_credential_env_keys(record.get("credential_env_keys", []))
        if _credential_error(keys):
            problems.append({"code": "credential_plaintext", "source_id": source_id})
        uri = str(record.get("uri") or "")
        if record.get("source_type") not in {"wecom_smartsheet"} and _uri_has_sensitive_query(uri):
            problems.append({"code": "unsafe_source_uri", "source_id": source_id})
        if record.get("source_type") in LOCAL_SOURCE_TYPES:
            allowed_root = str(record.get("allowed_root") or RAW_DIR)
            try:
                target, root = _validate_local_path(str(record.get("source_type")), uri, allowed_root)
                if Path(target).exists() and Path(root).exists() and not _path_is_under(Path(target), Path(root)):
                    problems.append({"code": "path_outside_allowed_root", "source_id": source_id})
            except PermissionError:
                problems.append({"code": "path_outside_allowed_root", "source_id": source_id})
            except Exception as exc:
                problems.append({"code": "local_path_invalid", "source_id": source_id, "message": str(exc)})
        if record.get("status") == "failed":
            problems.append({"code": "failed_source", "source_id": source_id})
    snapshots = load_snapshots()
    invalid_snapshots = [item for item in snapshots if not item.get("source_id") or not item.get("snapshot_id")]
    if invalid_snapshots:
        problems.append({"code": "invalid_snapshot_manifest", "count": len(invalid_snapshots)})
    status = "fail" if any(item["code"] in {"credential_plaintext", "unsafe_source_uri", "path_outside_allowed_root", "registry_unreadable"} for item in problems) else "warn" if problems else "ok"
    return _json(
        {
            "status": status,
            "schema": SOURCE_REGISTRY_SCHEMA,
            "source_count": len(sources),
            "snapshot_count": len(snapshots),
            "problem_count": len(problems),
            "problems": problems,
            "registry_path": str(SOURCE_REGISTRY_PATH),
            "snapshot_manifest_path": str(SNAPSHOT_MANIFEST_PATH),
        }
    )


def _main() -> None:
    parser = argparse.ArgumentParser(description="A2A Source Registry tools")
    sub = parser.add_subparsers(dest="command", required=True)
    register = sub.add_parser("register-source")
    register.add_argument("--source-id", required=True)
    register.add_argument("--display-name", required=True)
    register.add_argument("--source-type", required=True)
    register.add_argument("--uri", default="")
    register.add_argument("--allowed-root", default="")
    register.add_argument("--sync-mode", default="on_demand")
    register.add_argument("--owner", default="")
    register.add_argument("--sensitivity-level", default="internal")
    register.add_argument("--freshness-sla", default="")
    register.add_argument("--status", default="active")
    register.add_argument("--credential-env-keys", default="")
    register.add_argument("--metadata-json", default="{}")
    status = sub.add_parser("set-status")
    status.add_argument("source_id")
    status.add_argument("status")
    rebind = sub.add_parser("rebind-source")
    rebind.add_argument("source_id")
    rebind.add_argument("--uri", default="")
    rebind.add_argument("--allowed-root", default="")
    rebind.add_argument("--metadata-json", default="{}")
    sub.add_parser("list-sources")
    show = sub.add_parser("get-source")
    show.add_argument("source_id")
    sync = sub.add_parser("sync-source")
    sync.add_argument("source_id")
    sync.add_argument("--requested-by", default="frontend")
    workflow = sub.add_parser("run-workflow")
    workflow.add_argument("source_id")
    workflow.add_argument("--requested-by", default="frontend")
    health = sub.add_parser("health")
    health.set_defaults(_health=True)
    args = parser.parse_args()
    if args.command == "register-source":
        try:
            metadata = json.loads(args.metadata_json or "{}")
        except json.JSONDecodeError as exc:
            print(_error("invalid_metadata_json", str(exc)))
            return
        print(
            register_source(
                source_id=args.source_id,
                display_name=args.display_name,
                source_type=args.source_type,
                uri=args.uri,
                allowed_root=args.allowed_root,
                sync_mode=args.sync_mode,
                owner=args.owner,
                sensitivity_level=args.sensitivity_level,
                credential_env_keys=args.credential_env_keys,
                freshness_sla=args.freshness_sla,
                status=args.status,
                metadata=metadata if isinstance(metadata, dict) else {},
            )
        )
    elif args.command == "set-status":
        print(set_source_status(args.source_id, args.status))
    elif args.command == "rebind-source":
        try:
            metadata = json.loads(args.metadata_json or "{}")
        except json.JSONDecodeError as exc:
            print(_error("invalid_metadata_json", str(exc)))
            return
        print(
            rebind_source_path(
                args.source_id,
                uri=args.uri,
                allowed_root=args.allowed_root,
                metadata=metadata if isinstance(metadata, dict) else {},
            )
        )
    elif args.command == "list-sources":
        print(list_sources())
    elif args.command == "get-source":
        print(get_source(args.source_id))
    elif args.command == "sync-source":
        print(sync_source(args.source_id, requested_by=args.requested_by))
    elif args.command == "run-workflow":
        print(run_source_sync_workflow(args.source_id, requested_by=args.requested_by))
    elif args.command == "health":
        print(check_source_registry_health())


if __name__ == "__main__":
    _main()
